# coding=utf-8
import logging, traceback, pyodbc
from io import StringIO
from datetime import datetime
from xml.dom import minidom

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.writer.excel import save_virtual_workbook

from . import constants, models, utils
from app.models import AccountRenew
from .dbhelper import initdb
from .utils import UIException

logger = logging.getLogger(__name__)

def authentication_list(batch_id, dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
            
        q = """
            select cid.batchid, cid.callerid, cid.assigndate, cid.wskey, cid.code_area, cid.status from gw_BatchNumber bn, GW_callerid cid
            where bn.batch_id=cid.batchid and bn.batch_id like ? and  cid.status =0 

            """
           
        rows = db.cur.execute(q,'%{0}%'.format(batch_id)).fetchall()
        for r in rows:
            o = models.callerid_detail()
            o.callerid = r.callerid
            o.batchid = r.batchid
            o.status = r.status
            o.assigndate =r.assigndate 
            o.wskey = r.wskey
            o.code_area = r.code_area
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def batch_list(batch_id, dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
            
        q = """
             select batch_id, batch_date, uploadby from gw_BatchNumber 
            where batch_id = ?
            """
        rows = db.cur.execute(q, batch_id).fetchall()
        for r in rows:
            o = models.BatchDetail()
            o.batch_id = r.batch_id
            o.batch_date = r.batch_date
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def authentication_delete(idxlist, user, pwd):
    db = None
    
    try:
        if idxlist is None:
            raise UIException('No selected callerid')
        
        if len(idxlist) < 1:
            raise UIException('No selected callerid')
        
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        
        params = ','.join(["'{0}'".format(i) for i in idxlist])
        q = 'delete from authentication where callerid in ({0})'.format(params)
        db.cur.execute(q)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def authentication_suspend(idxlist, accountid, status, user, pwd):
    db = None
    
    try:
        if idxlist is None:
            raise UIException('No selected callerid')
        
        if len(idxlist) < 1:
            raise UIException('No selected callerid')
        
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        
        params = ','.join(["'{0}'".format(i) for i in idxlist])
        q = 'update authentication set status = ? where accountid = ? and callerid in ({0})'.format(params)
        db.cur.execute(q, status, accountid)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def authentication_suspendall(accountid, status, user, pwd):
    db = None
    
    try:
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        
        q = 'update authentication set status = ? where accountid = ?'
        db.cur.execute(q, status, accountid)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def authentication_add(o, user, pwd):
    b = False
    db = None
    
    try:
        assert isinstance(o, models.Authentication)
        
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        
        authentication_create(db, o)
        db.commit()
            
    finally:
        if db is not None:
            db.dispose()
            
    return b

def authentication_update(oldcallerid, o, user, pwd):
    db = None
    
    try:
        assert isinstance(o, models.Authentication)
        
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        
        q = 'select accountid from authentication where callerid = ?'
        row = db.cur.execute(q, o.callerid).fetchone()
        if row is not None:
            raise UIException('CallerID {0} already exist in the database'.format(o.callerid))
        
        q = 'delete from authentication where callerid = ?'
        db.cur.execute(q, oldcallerid)
        
        q = """
            insert into authentication (callerid, status, accountid) 
            values (?, ?, ?)
            """
        db.cur.execute(q, o.callerid, o.status, o.accountid)
        db.commit()
    
    finally:
        if db is not None:
            db.dispose()

def authentication_bulkcreate(startcallerid, qty):
    l = []
    
    if qty == 0:
        raise UIException('No callerid to generate')
    
    k = startcallerid
    for i in range(qty):
        x = int(k) + i
        s = str(x).zfill(len(startcallerid))
        l.append('{0},0'.format(s))
        
    return '\n'.join(l)

def authentication_bulksave(data, accountid, user, pwd):
    b = False
    db = None
    
    try:
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        
        l = data.split('\n')
        delimiter = ','
        n = 0
        
        for s in l:
            a = s.replace(';', delimiter).replace('|', delimiter)
            k = a.split(delimiter)
            
            o = models.Authentication()
            o.accountid = accountid
            o.callerid = k[0].strip()
            o.status = 1 if len(k) < 2 else k[1]
            o.status = 1 if o.status != 0 else 0
            
            authentication_create(db, o)
            n += 1
            
        if n == len(l):
            db.commit()
            b = True
        
    finally:
        if db is not None:
            db.dispose()
            
    return b, n

def authentication_create(db, o):
    try:
        assert isinstance(o, models.Authentication)
        
        q = """
            insert into authentication (callerid, status, accountid) 
            values (?, ?, ?)
            """
        db.cur.execute(q, o.callerid, o.status, o.accountid)
    
    except pyodbc.IntegrityError as e:
        if utils.isduplicatekey(e):
            raise UIException('CallerID {0} already exist'.format(o.callerid))
        
        else:
            raise UIException(traceback.format_exc())

def authentication_search(data):
    lr = []
    db = None

    try:
        db = initdb()

        l = data.split('\n')
        lk = []
        m = {}
        i = 0
        for s in l:
            v = s.strip()
            lk.append("'{0}'".format(v))
            m[v] = i
            lr.append({
                'callerid': v,
                'accountid': '',
                'name': '',
                'status': ''
            })
            i += 1

        k = ','.join(lk)
        d = models.ACSTATUS_DIC
        q = """
            select a.callerid, b.accountid, b.name, c.acstatus from authentication a 
            inner join userdetail b on a.accountid = b.accountid 
            inner join userdetailext c on b.accountid = c.accountid 
            where a.callerid in ({0})
            """.format(k)
        print(q)
        rows = db.cur.execute(q).fetchall()
        for r in rows:
            if r.callerid in m:
                j = m[r.callerid]
                status = ''
                if r.acstatus in d:
                    status = d[r.acstatus]

                lr[j] = {
                    'callerid': r.callerid,
                    'accountid': r.accountid,
                    'name': r.name,
                    'status': status
                }

    finally:
        if db is not None:
            db.dispose()

    return lr

def bbauthentication_list_available(batch_id, dbx=None):
    l = []
    db = dbx
    # print("running here")
    try:
        db = initdb(dbx)
            
        q = """
            select cid.batchid, cid.callerid, cid.assigndate, cid.assignee, cid.wskey, cid.code_area, cid.status, cid.releasedate, cid.blockdate, cid.state
            from gw_BatchNumber bn, GW_callerid cid
            where bn.batch_id=cid.batchid and bn.batch_id like ? and  cid.status =0 
            """

            #left join MaxCallAppearanceDN d on a.callerid = d.DirectoryNumber


        rows = db.cur.execute(q,'%{0}%'.format(batch_id)).fetchall()
        for r in rows:
            o = models.callerid_detail()
            o.callerid = r.callerid
            o.batchid = r.batchid
            o.status = r.status
            o.assigndate =r.assigndate 
            o.releasedate =r.releasedate 
            o.blockdate =r.blockdate 
            o.state =r.state 
            o.wskey = r.wskey
            o.assignee = r.assignee
            o.code_area = r.code_area
            l.append(o)

         
            l.append(o)
            #print(r.callerid)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l


def bbauthentication_list_blocked(batch_id, dbx=None):
    l = []
    db = dbx
    # print("running here")
    try:
        db = initdb(dbx)
            
        q = """
            select cid.batchid, cid.callerid, cid.assigndate, cid.assignee, cid.wskey, cid.code_area, cid.status, cid.releasedate, cid.blockdate, cid.state
            from gw_BatchNumber bn, GW_callerid cid
            where bn.batch_id=cid.batchid and bn.batch_id like ? and  cid.status =2 
            """

            #left join MaxCallAppearanceDN d on a.callerid = d.DirectoryNumber


        rows = db.cur.execute(q,'%{0}%'.format(batch_id)).fetchall()
        for r in rows:
            o = models.callerid_detail()
            o.callerid = r.callerid
            o.batchid = r.batchid
            o.status = r.status
            o.assigndate =r.assigndate 
            o.releasedate =r.releasedate 
            o.blockdate =r.blockdate 
            o.state =r.state 
            o.wskey = r.wskey
            o.code_area = r.code_area
            o.assignee = r.assignee

            l.append(o)

         
            l.append(o)
            #print(r.callerid)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l


def bbauthentication_list_used(batch_id, dbx=None):
    l = []
    db = dbx
    # print("running here")
    try:
        db = initdb(dbx)
            
        q = """
            select cid.batchid, cid.callerid, cid.assigndate, cid.assignee, cid.wskey, cid.code_area, cid.status, cid.releasedate, cid.blockdate, cid.state
            from gw_BatchNumber bn, GW_callerid cid
            where bn.batch_id=cid.batchid and bn.batch_id like ? and  cid.status =1 
            """

            #left join MaxCallAppearanceDN d on a.callerid = d.DirectoryNumber


        rows = db.cur.execute(q,'%{0}%'.format(batch_id)).fetchall()
        for r in rows:
            o = models.callerid_detail()
            o.callerid = r.callerid
            o.batchid = r.batchid
            o.status = r.status
            o.assigndate =r.assigndate 
            o.releasedate =r.releasedate 
            o.blockdate =r.blockdate 
            o.state =r.state 
            o.wskey = r.wskey
            o.code_area = r.code_area
            o.assignee = r.assignee

            l.append(o)

         
            l.append(o)
            #print(r.callerid)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l


def bbauthentication_list_testing(batch_id, dbx=None):
    l = []
    db = dbx
    # print("running here")
    # 0 available
    # 1used
    # 2 block
    #3 test
    try:
        db = initdb(dbx)
            
        q = """
            select cid.batchid, cid.callerid, cid.assigndate, cid.assignee, cid.wskey, cid.code_area, cid.status, cid.releasedate, cid.blockdate, cid.state
            from gw_BatchNumber bn, GW_callerid cid
            where bn.batch_id=cid.batchid and bn.batch_id like ? and  cid.status =3 
            """

            #left join MaxCallAppearanceDN d on a.callerid = d.DirectoryNumber


        rows = db.cur.execute(q,'%{0}%'.format(batch_id)).fetchall()
        for r in rows:
            o = models.callerid_detail()
            o.callerid = r.callerid
            o.batchid = r.batchid
            o.status = r.status
            o.assigndate =r.assigndate 
            o.releasedate =r.releasedate 
            o.blockdate =r.blockdate 
            o.state =r.state 
            o.wskey = r.wskey
            o.code_area = r.code_area
            o.assignee = r.assignee

            l.append(o)

         
            l.append(o)
            #print(r.callerid)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l


def bbauthentication_list_add(batch_id, dbx=None):
    l = []
    db = dbx
    # print("running here")
    try:
        db = initdb(dbx)
            
        q = """
            select cid.batchid, cid.callerid, cid.assigndate, cid.assignee, cid.wskey, cid.code_area, cid.status, cid.releasedate, cid.blockdate, cid.state
            from gw_BatchNumber bn, GW_callerid cid
            where bn.batch_id=cid.batchid and bn.batch_id like ? and  cid.status =0 
            """

            #left join MaxCallAppearanceDN d on a.callerid = d.DirectoryNumber


        rows = db.cur.execute(q,'%{0}%'.format(batch_id)).fetchall()
        for r in rows:
            o = models.callerid_detail()
            o.callerid = r.callerid
            o.batchid = r.batchid
            o.status = r.status
            o.assigndate =r.assigndate 
            o.releasedate =r.releasedate 
            o.blockdate =r.blockdate 
            o.state =r.state 
            o.wskey = r.wskey
            o.code_area = r.code_area
            o.assignee = r.assignee

            l.append(o)

         
            l.append(o)
            #print(r.callerid)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l


def bbauthentication_list(accountid, dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)

        q = """
            select a.accountid, a.callerid, b.bb_rt015, b.bb_015pwd, b.bb_allowpstn, b.bb_displayname, a.status, 
            b.bb_prepaid, b.bb_forward, c.subnetmask, d.MaxCallAppearance, d.DirectoryNumber, ud.igatetype, ud.pbxno
            from authentication a inner join bb_authentication b on a.callerid = b.bb_rt015 
            left join MaxCallAppearanceDN d on a.callerid = d.DirectoryNumber
			left join userdetail ud on ud.pbxno = a.callerid
            left join openquery (SIPSERVER, 'select * from authsrcip') c on a.callerid = c.username 
            where a.accountid = ?
            order by a.callerid
            """

        rows = db.cur.execute(q, accountid).fetchall()
        for r in rows:
            o = models.BBAuthentication()
            o.bb_rt015 = r.bb_rt015
            o.bb_015pwd = r.bb_015pwd
            o.bb_allowpstn = r.bb_allowpstn
            o.bb_displayname = r.bb_displayname
            o.bb_prepaid = r.bb_prepaid
            o.bb_forward = r.bb_forward
            o.subnetmask = r.subnetmask
            
            x = models.Authentication()
            x.callerid = r.callerid
            x.status = r.status
            x.accountid = r.accountid

            ud = models.UserDetail()
            ud.accountid = r.accountid
            ud.pbxno = r.pbxno
            ud.igatetype = r.igatetype

            y = models.MaxCallAppearanceDN()
            y.DirectoryNumber = r.DirectoryNumber
            y.MaxCallAppearance = r.MaxCallAppearance
            
            o.maxcallappearancedn = y
            o.authentication = x
            o.UserDetail=ud

            # calling api
            # if r.igatetype == 7:
            #     u = ('https://billing.redtone.com/metaswitchapi1/api/pbx/{0}'.format(ud.pbxno))
            #     k = requests.get(u, verify=False)
            #     j = k.json()
            #     # print("==get=="+str(j))
            #     if j.get('success') == 1:
            #         print("Success")
            #     else:
            #         print("Pending")
            

            # if r.igatetype == 6:
            #     print(r.callerid)
            #     u = ('https://billing.redtone.com/metaswitchapi1/api/subs/{0}'.format(r.callerid))
            #     k = requests.get(u, verify=False)
            #     j = k.json()
            #     # print("==get=="+str(j))
            #     if j.get('success') == 1:
            #         print("Success")
            #     else:
            #         print("Pending")


         
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l


def bbauthentication_listws(wholesalerkey, nicenum=False, search=None, dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
        
        # prevent selection for existing prepaid numbers
        q = """
            select bb_rt015, bb_015pwd, bb_allowpstn, bb_nicenum, bb_wskey, bb_deregisterdate 
            from bb_authentication where bb_status = 0 and bb_wskey = ?
            and bb_rt015 not in (
                select username from openquery(sipserver, 'select username, grp from grp')
                where grp = 'prepaid'
            )
            """
        lq = [q]
        prm = [wholesalerkey]
        
        if nicenum:
            lq.append(' and bb_nicenum = 1')
            
        if search is not None:
            lq.append(' and bb_rt015 like ?')
            prm.append('%{0}%'.format(search))
            
        lq.append(' order by bb_rt015')
        params = tuple(prm)
        
        q = ''.join(lq)
        rows = db.cur.execute(q, params).fetchall()
        for r in rows:
            o = models.BBAuthentication()
            o.bb_rt015 = r.bb_rt015
            o.bb_015pwd = r.bb_015pwd
            o.bb_allowpstn = r.bb_allowpstn
            o.bb_nicenum = r.bb_nicenum
            o.bb_wskey = r.bb_wskey
            o.bb_deregisterdate = r.bb_deregisterdate
            l.append(o)
            
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def bbauthentication_listws_excel_list(wholesalerkey, nicenum=False, search=None):
    b = None
    
    try:
        wb = Workbook()
        ws = wb.active
        
        l = bbauthentication_listws(wholesalerkey, nicenum, search)
        
        ws.column_dimensions['A'].width = 18
        lh = ['RT 015/03 Number', 'PSTN', 'Nice No']
        j = 1
        for h in lh:
            ws.cell(row=1, column=j).value = h
            j += 1
            
        i = 2
        for x in l:
            ws.cell(row=i, column=1).value = x.bb_rt015
            ws.cell(row=i, column=2).value = x.bb_allowpstn
            ws.cell(row=i, column=3).value = x.bb_nicenum
            i += 1
            
        b = save_virtual_workbook(wb)
        
    except:
        raise
            
    return b

def bbauthentication_delete(idxlist, accountid, user, pwd ):
    le = []
    db = None
    
    try:
        if idxlist is None:
            raise UIException('No selected number')
        
        if len(idxlist) < 1:
            raise UIException('No selected number')
        
        db = initdb(user=user, pwd=pwd)
        
        for dic in idxlist:
            ex = bbauthentication_suspend_callerid(db, dic['callerid'], suspend=2)
            print(dic)
            if ex is not None:
                le.append(ex)
        
    finally:
        if db is not None:
            db.dispose()
            
    return le


# def bbauthentication_delete_number(idxlist, accountid, user, pwd ):
#     le = []
#     db = None
    
#     try:
#         if idxlist is None:
#             raise UIException('No selected number')
        
#         if len(idxlist) < 1:
#             raise UIException('No selected number')
        
#         db = initdb(user=user, pwd=pwd)
        
#         for dic in idxlist:
#             ex = bbauthentication_move_callerid(db, dic['callerid'])
#             print(dic)
#             if ex is not None:
#                 le.append(ex)
        
#     finally:
#         if db is not None:
#             db.dispose()
            
#     return le
def bbauthentication_delete_number(idxlist, accountid, user, pwd):
    le = []
    db = None
    
    try:
        if idxlist is None:
            raise UIException('No selected number')
        
        if len(idxlist) < 1:
            raise UIException('No selected number')
        
        db = initdb(user=user, pwd=pwd)
        
        for dic in idxlist:
            ex = bbauthentication_move_callerid(db, dic['callerid'], accountid,user, pwd)
            print(dic)

            if ex is not None:
                le.append(ex)
                
    finally:
        if db is not None:
            db.dispose()
        
    return le


def bbauthentication_block(idxlist, accountid, user, pwd):
    le = []
    db = None
    
    try:
        if idxlist is None:
            raise UIException('No selected number')
        
        if len(idxlist) < 1:
            raise UIException('No selected number')
        
        db = initdb(user=user, pwd=pwd)
        
        for dic in idxlist:
            ex = bbauthentication_block_callerid(db, dic['callerid'], accountid,user, pwd)
            print(dic)

            if ex is not None:
                le.append(ex)
                
    finally:
        if db is not None:
            db.dispose()
        
    return le




def bbauthentication_resume(idxlist, accountid, user, pwd):
    le = []
    db = None
    
    try:
        if idxlist is None:
            raise UIException('No selected number')
        
        if len(idxlist) < 1:
            raise UIException('No selected number')
        
        db = initdb(user=user, pwd=pwd)
        
        for dic in idxlist:
            ex = bbauthentication_resume_callerid(db, dic['callerid'], accountid,user, pwd)
            print(dic)

            if ex is not None:
                le.append(ex)
                
    finally:
        if db is not None:
            db.dispose()
        
    return le

def bbauthentication_resumetestnum(idxlist, accountid, user, pwd):
    le = []
    db = None
    
    try:
        if idxlist is None:
            raise UIException('No selected number')
        
        if len(idxlist) < 1:
            raise UIException('No selected number')
        
        db = initdb(user=user, pwd=pwd)
        
        for dic in idxlist:
            ex = bbauthentication_resume_callerid_testnum(db, dic['callerid'], accountid,user, pwd)
            print(dic)

            if ex is not None:
                le.append(ex)
                
    finally:
        if db is not None:
            db.dispose()
        
    return le


def bbauthentication_releasenumber(idxlist, accountid, user, pwd):
    le = []
    db = None
    
    try:
        if idxlist is None:
            raise UIException('No selected number')
        
        if len(idxlist) < 1:
            raise UIException('No selected number')
        
        db = initdb(user=user, pwd=pwd)
        
        for dic in idxlist:
            ex = bbauthentication_releasenumber_callerid(db, dic['callerid'], accountid,user, pwd)
            print(dic)

            if ex is not None:
                le.append(ex)
                
    finally:
        if db is not None:
            db.dispose()
        
    return le


def bbauthentication_settestnum(idxlist, accountid, user, pwd):
    le = []
    db = None
    
    try:
        if idxlist is None:
            raise UIException('No selected number')
        
        if len(idxlist) < 1:
            raise UIException('No selected number')
        
        db = initdb(user=user, pwd=pwd)
        
        for dic in idxlist:
            ex = bbauthentication_settest_callerid(db, dic['callerid'], accountid,user, pwd)
            print(dic)

            if ex is not None:
                le.append(ex)
                
    finally:
        if db is not None:
            db.dispose()
        
    return le


def bbauthentication_changepwd(idxlist, accountid, username):
    le = []
    db = None
    
    try:
        if idxlist is None:
            raise UIException('No selected number')
        
        if len(idxlist) < 1:
            raise UIException('No selected number')
        
        db = initdb()
        
        for callerid in idxlist:
            password = utils.gensippassword()
            ex = bbauthentication_changepwd_callerid(db, callerid, password, accountid, username)
            if ex is not None:
                le.append(ex)
                
    finally:
        if db is not None:
            db.dispose()
            
    return le

def bbauthentication_pstntoggle(idxlist, user, pwd):
    db = None
    
    try:
        if idxlist is None:
            raise UIException('No selected number')
        
        if len(idxlist) < 1:
            raise UIException('No selected number')
        
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        
        params = ','.join(["'{0}'".format(i) for i in idxlist])
        q = 'update bb_authentication set bb_allowpstn = (bb_allowpstn + 1) % 2 where bb_rt015 in ({0})'.format(params)
        db.cur.execute(q)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def bbauthentication_pstnset(idxlist, pstn, user, pwd):
    db = None
    
    try:
        if idxlist is None:
            raise UIException('No selected number')
        
        if len(idxlist) < 1:
            raise UIException('No selected number')
        
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        
        params = ','.join(["'{0}'".format(i) for i in idxlist])
        q = 'update bb_authentication set bb_allowpstn = ? where bb_rt015 in ({0})'.format(params)
        db.cur.execute(q, pstn)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def bbauthentication_nicenumtoggle(idxlist, user, pwd):
    db = None
    
    try:
        if idxlist is None:
            raise UIException('No selected number')
        
        if len(idxlist) < 1:
            raise UIException('No selected number')
        
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        
        params = ','.join(["'{0}'".format(i) for i in idxlist])
        q = 'update bb_authentication set bb_nicenum = (bb_nicenum + 1) % 2 where bb_rt015 in ({0})'.format(params)
        db.cur.execute(q)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def bbauthentication_nicenumset(idxlist, nicenum, user, pwd):
    db = None
    
    try:
        if idxlist is None:
            raise UIException('No selected number')
        
        if len(idxlist) < 1:
            raise UIException('No selected number')
        
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        
        params = ','.join(["'{0}'".format(i) for i in idxlist])
        q = 'update bb_authentication set bb_nicenum = ? where bb_rt015 in ({0})'.format(params)
        db.cur.execute(q, nicenum)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def bbauthentication_generatepwd(idxlist, user, pwd):
    db = None
    
    try:
        if idxlist is None:
            raise UIException('No selected number')
        
        if len(idxlist) < 1:
            raise UIException('No selected number')
        
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        
        for callerid in idxlist:
            password = utils.gensippassword()
            q = 'update bb_authentication set bb_015pwd = ? where bb_rt015 = ?'
            db.cur.execute(q, password, callerid)
        
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def bbauthentication_suspend_callerid(dbx, callerid, suspend=2):
    ex = None
    db = dbx
    
    try:
        db = initdb(dbx)

        insert_into_blocked_record(batchid, callerid, wskey, code_area,status, user)
        
        q = """
            update GW_callerid set status = ? where callerid = ? 
            """
        print(callerid)
        db.cur.execute(q, suspend, callerid)
        db.commit()


    except:
        ex = traceback.format_exc()
        
    finally:
        if db is not None:
            db.dispose()
            
    return ex


def bbauthentication_move_callerid(dbx, callerid, accountid,user, pwd):
    ex = None
    db = None
    print("delete number")
    try:
        # db = initdb(dbx)
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)

        q = """
            select cid.callerid,cid.batchid, cid.wskey, cid.state, cid.code_area, cid.status from gw_BatchNumber bn, GW_callerid cid
            where bn.batch_id=cid.batchid and cid.callerid like ? 
            """
        r = db.cur.execute(q, callerid).fetchone()
        if r is not None:
            callerid = r.callerid
            batchid = r.batchid
            wskey = r.wskey
            state = r.state
            status = r.status
            code_area = r.code_area

        print(batchid)

        #is3
        print(user)
        insert_into_deleted_record(batchid, callerid, wskey, code_area,status, user)


        q = """
            DELETE FROM GW_callerid where callerid = ? 
            """
        print(callerid)
        db.cur.execute(q,  callerid)
        db.commit()

        
    except:
        ex = traceback.format_exc()
        
    finally:
        if db is not None:
            db.dispose()
            
    return ex




def bbauthentication_delete_callerid(dbx, callerid, suspend=2):
    ex = None
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = """
            update GW_callerid set status = ? where callerid = ? 
            """
        print(callerid)
        db.cur.execute(q, suspend, callerid)
        db.commit()


    except:
        ex = traceback.format_exc()
        
    finally:
        if db is not None:
            db.dispose()
            
    return ex



#start added by ainur
def update_status_wscallbilling(status, accountid):
    #db = dbx
    status=1
    # accountid='110050001'
    
    try:
        
        # if status == 1:
           # q = """
            #update UserDetailExt_dup set AcStatus=? where AccountID=?"
            #    """
            #update Account_MasterData set Status=? where AccountID=?"

            #r = db.cur.execute(q, status, accountid).fetchone()

    #update dbo.Customer set [Status]=@AccountStatus,[LastUpdateInfo]=GETDATE() where [CustID]=@CustID
	#update dbo.Account_MasterData set [Status]=@AccountStatus where AccountID=@CustID  
       db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        
       q = """
        update UserDetailExt_dup set AcStatus=? where AccountID=?"

        """
       db.cur.execute(q, status, accountid)
       db.commit()
        
    finally:
        if db is not None:
            db.dispose()


#end added

def bbauthentication_block_callerid(dbx, callerid, accountid,user, pwd):
    ex = None
    db = None
    print(accountid)
    try:
        # db = initdb(dbx)
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        q = """
            select cid.callerid,cid.batchid, cid.wskey, cid.state, cid.code_area, cid.status from gw_BatchNumber bn, GW_callerid cid
            where bn.batch_id=cid.batchid and cid.callerid like ? 
            """
        r = db.cur.execute(q, callerid).fetchone()
        if r is not None:
            callerid = r.callerid
            batchid = r.batchid
            wskey = r.wskey
            state = r.state
            status = r.status
            code_area = r.code_area

        print(batchid)

        insert_into_blocked_record(batchid, callerid, wskey, code_area,2, user)

        q = """
            update GW_callerid set status = 2, blockdate=? where callerid = ? 
            """
        print(callerid)
        db.cur.execute(q, datetime.now(), callerid)
        db.commit()

        
    except:
        ex = traceback.format_exc()
        
    finally:
        if db is not None:
            db.dispose()
            
    return ex



def bbauthentication_resume_callerid(dbx, callerid, accountid,user, pwd):
    ex = None
    db = None
    print("set number as testing")
    try:
        # db = initdb(dbx)
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)


        q = """
            select cid.callerid,cid.batchid, cid.wskey, cid.state, cid.code_area, cid.status from gw_BatchNumber bn, GW_callerid cid
            where bn.batch_id=cid.batchid and cid.callerid like ? 
            """
        r = db.cur.execute(q, callerid).fetchone()
        if r is not None:
            callerid = r.callerid
            batchid = r.batchid
            wskey = r.wskey
            state = r.state
            status = r.status
            code_area = r.code_area

        print(batchid)

        insert_into_blocked_record(batchid, callerid, wskey, code_area,0, user)


        q = """
            update GW_callerid set status = 0, releasedate=? where callerid = ? 
            """
        print(callerid)
        db.cur.execute(q, datetime.now(),callerid)
        db.commit()

        
    except:
        ex = traceback.format_exc()
        
    finally:
        if db is not None:
            db.dispose()
            
    return ex

# def bbauthentication_resume_callerid_testnum(dbx, callerid, accountid,user, pwd):
#     ex = None
#     db = None
#     print("set number as testing")
#     try:
#         # db = initdb(dbx)
#         db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)


#         q = """
#             update GW_callerid set status = 0, releasedate=? where callerid = ? 
#             """
#         print(callerid)
#         db.cur.execute(q,datetime.now(), callerid)
#         db.commit()

        
#     except:
#         ex = traceback.format_exc()
        
#     finally:
#         if db is not None:
#             db.dispose()
            
#     return ex


# def bbauthentication_releasenumber_callerid(dbx, callerid, accountid,user, pwd):
#     ex = None
#     db = None
#     print("set number as testing")
#     try:
#         # db = initdb(dbx)
#         db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)

#         q = """
#             update GW_callerid set status = 0,releasedate=? where callerid = ? 
#             """
#         print(callerid)
#         db.cur.execute(q, datetime.now(), callerid)
#         db.commit()

        
#     except:
#         ex = traceback.format_exc()
        
#     finally:
#         if db is not None:
#             db.dispose()
            
#     return ex


def bbauthentication_settest_callerid(dbx, callerid, accountid,user, pwd):
    ex = None
    db = None
    print("set number as testing")
    try:
        # db = initdb(dbx)
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)

        # q = """
        #     SET NOCOUNT ON
        #     declare
        #       @return_value int,
        #       @ErrorMsg varchar(2000)
        #     exec @return_value = spp_Register015
        #       @CallerID = ?,
        #       @AccountID = ?,
        #       @Password = ?,
        #       @AllowPSTN = ?,
        #       @ErrorMsg = @ErrorMsg output
        #     select
        #       @ErrorMsg as N'ErrorMsg',
        #       @return_value as N'Ret'
        #     """
        # q = """update GW_callerid set status = 2 where callerid = ? and batchid = ? """

           
        # r = db.cur.execute(q, callerid, accountid).fetchone()
        # print(r)
        # if r.Ret != 1:
        #     raise Exception('Failed to register sip number {0}'.format(callerid))

        q = """
            update GW_callerid set status = 3 where callerid = ? 
            """
        print(callerid)
        db.cur.execute(q, callerid)
        db.commit()

        
    except:
        ex = traceback.format_exc()
        
    finally:
        if db is not None:
            db.dispose()
            
    return ex


def bbauthentication_update_displayname(name, callerid, user, pwd):
    db = None
    
    try:
        v = name.strip()
        
        if v.isdigit() == False:
            raise UIException('Invalid display name {0}, must be number only'.format(v))
        
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        
        q = 'update bb_authentication set bb_displayname = ? where bb_rt015 = ?'
        db.cur.execute(q, v, callerid)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def bbauthentication_update_MCA(MaxCallAppearance, callerid, user, pwd):
    db = None
    
    try:
        v = MaxCallAppearance.strip()
        
        if v.isdigit() == False:
            raise UIException('Invalid value {0}, must be number only'.format(v))
        
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        print(callerid)
        print(MaxCallAppearance)
        #q = 'update MaxCallAppearanceDN set MaxCallAppearance = ? where DirectoryNumber = ?'
        #db.cur.execute(q, v, callerid)

        q = """
            if not exists(select DirectoryNumber from MaxCallAppearanceDN where DirectoryNumber = ?)
        begin
            insert into MaxCallAppearanceDN (DirectoryNumber, MaxCallAppearance )
            values(?,?)
        end
        else
        begin
            update MaxCallAppearanceDN set MaxCallAppearance = ? where DirectoryNumber = ?
        end
        """ 
        db.cur.execute(q, callerid, callerid, MaxCallAppearance, MaxCallAppearance, callerid)

        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def bbauthentication_setcallforward(callerid, callfwd, accountid, username):
    db = None
    
    try:
        db = initdb()
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int
            exec @return_value = spp_SetCallForward
              @SipNo = ?,
              @CallForward = ?
            select
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, callerid, callfwd.strip()).fetchone()
        if r.Ret != 1:
            raise Exception('Failed to set call forward')
        
        qs ='Set Sip Call forwarding {0}->{1}'.format(callerid, callfwd.strip())
        insert_into_sql_log(qs, accountid, username, db, -1)
            
    finally:
        if db is not None:
            db.dispose()

def bbauthentication_removecallforward(num, accountid, username):
    db = None
    
    try:
        db = initdb()
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int
            exec @return_value = spp_RemoveCallForward
              @SipNo = ?
            select
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, num.strip()).fetchone()
        if r.Ret != 1:
            raise Exception('Failed to remove call forward')
        
        qs = 'Remove Sip Call forwarding: {0}'.format(num.strip())
        insert_into_sql_log(qs, accountid, username, db, -1)
            
    finally:
        if db is not None:
            db.dispose()

def bbauthentication_changepwd_callerid(dbx, callerid, pwd, accountid, username):
    ex = None
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int,
              @ErrorMsg varchar(500)
            exec @return_value = spp_015ChangePassword
              @CallerID = ?,
              @NewPassword = ?,
              @ErrorMsg = @ErrorMsg output
            select
              @ErrorMsg as N'ErrorMsg',
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, callerid, pwd.strip()).fetchone()
        if r.Ret != 1:
            raise Exception('Failed to change sip number password {0}, Error: {1}'.format(callerid, r.ErrorMsg))
        
        else:
            qs = "UPDATE BB_Authentication SET bb_015pwd='{0}' WHERE bb_rt015='{1}'".format(pwd, callerid)
            insert_into_sql_log(qs, accountid, username, db, -1)
        
    except:
        ex = traceback.format_exc()
        
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return ex

def bbauthentication_update_subnetmask(callerid, subnetmask):
    db = None
    
    try:
        db = initdb()
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int,
              @ErrorMsg varchar(500)
            exec @return_value = spp_015SubnetMask_Update
              @CallerID = ?,
              @strSubnetMask = ?,
              @ErrorMsg = @ErrorMsg output
            select
              @ErrorMsg as N'ErrorMsg',
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, callerid, subnetmask.strip()).fetchone()
        if r.Ret != 1:
            raise Exception('Failed to update this sip number {0} subnetmask, Error: {1}'.format(callerid, r.ErrorMsg))
            
    finally:
        if db is not None:
            db.dispose()

def bbauthentication_update_sippstn(callerid, pstn, accountid, username):
    db = None
    
    try:
        db = initdb()
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int,
              @ErrorMsg varchar(500)
            exec @return_value = spp_015ChangeAllowPSTN
              @CallerID = ?,
              @AllowPSTN = ?,
              @ErrorMsg = @ErrorMsg output
            select 
              @ErrorMsg as N'ErrorMsg',
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, callerid, pstn).fetchone()
        if r.Ret != 1:
            raise Exception('Failed to update this sip number {0} AllowPSTN setting, Error: {1}'.format(callerid, r.ErrorMsg))
        
        qs = "UPDATE BB_Authentication SET bb_allowpstn={0} WHERE bb_rt015='{1}'".format(pstn, callerid)
        insert_into_sql_log(qs, accountid, username, db, -1)
            
    finally:
        if db is not None:
            db.dispose()

def bbauthentication_update_sipprepaid(callerid, prepaid, accountid, username):
    db = None
    
    try:
        db = initdb()
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int,
              @ErrorMsg varchar(500)
            exec @return_value = spp_015SetPrepaid
              @CallerID = ?,
              @Prepaid = ?,
              @ErrorMsg = @ErrorMsg output
            select
              @ErrorMsg as N'ErrorMsg',
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, callerid, prepaid).fetchone()
        if r.Ret != 1:
            raise Exception('Failed to set this sip number {0} to prepaid mode, Error: {1}'.format(callerid, r.ErrorMsg))
        
        qs = "UPDATE BB_Authentication SET bb_prepaid={0} WHERE bb_rt015='{1}'".format(prepaid, callerid)
        insert_into_sql_log(qs, accountid, username, db, -1)
        
    finally:
        if db is not None:
            db.dispose()
            
def pinuserid_list(accountid, dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = """
            select custid, pin, userid, description, creditlimit from pinuserid
            where custid = ?
            order by userid
            """
        rows = db.cur.execute(q, accountid).fetchall()
        for r in rows:
            o = models.PinUserID()
            o.custid = r.custid
            o.pin = r.pin
            o.userid = r.userid
            o.description = r.description
            o.creditlimit = r.creditlimit
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def pinuserid_importsave(pinlength, data, accountid, user, pwd, withdesc=False):
    b = False
    db = None
    
    try:
        if pinlength < 1:
            raise UIException('Invalid pin length')
        
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        
        l = data.split('\n')
        delimiter = '|'
        n = 0
        
        o = models.PinUserID()
        o.custid = accountid
        o.pin = ''
        o.userid = ''
        o.description = ''
        o.creditlimit = 0
        
        for s in l:
            if withdesc:
                a = s.replace(';', delimiter).replace(',', delimiter)
                k = a.split(delimiter)
                assert(len(k) == 4), 'Must be in this format pin|userid|creditlimit|description'
                
                o.pin = k[0]
                o.userid = k[1]
                o.creditlimit = float(k[2])
                o.description = k[3]
                
            else:
                o.pin = s[:pinlength].strip()
                o.userid = s[pinlength:]
                
            pinuserid_create(db, o)
            n += 1
                
        if n == len(l):
            db.commit()
            b = True
            
    finally:
        if db is not None:
            db.dispose()
            
    return b, n

def pinuserid_create(db, o):
    try:
        assert isinstance(o, models.PinUserID)
        
        lastuserid = o.userid
        
        if o.userid == '':
            lastuserid = pinuserid_lastuserid(o.custid, db)
            
        q = """
            insert into pinuserid (custid, pin, userid, description, creditlimit) 
            values (?, ?, ?, ?, ?)
            """
        db.cur.execute(q, o.custid, o.pin, lastuserid, o.description, o.creditlimit)
        
    except pyodbc.IntegrityError as e:
        if utils.isduplicatekey(e):
            raise UIException('Pin {0} already exist'.format(o.pin))
        
        else:
            raise UIException(traceback.format_exc())

def pinuserid_lastuserid(accountid, dbx=None):
    userid = 0
    i = 0
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = 'select max(userid) as maxx from pinuserid where custid = ?'
        r = db.cur.execute(q, accountid).fetchone()
        if r is not None:
            x = r.maxx
            i = int(x) if x is not None else 0
            
        userid = i + 1
        
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return '{0}'.format(userid).zfill(4)

def pinuserid_userid_update(accountid, pin, userid, newuserid, user, pwd):
    db = None
    
    try:
        if newuserid is None or newuserid == '':
            raise UIException('New UserID is required')
        
        if len(newuserid) > 4:
            raise UIException("UserID's maximum length is 4 digit")
        
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        
        q = """
            update pinuserid set userid = ?
            where custid = ? and pin = ? and userid = ?
            """
        db.cur.execute(q, newuserid, accountid, pin, userid)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def pinuserid_delete(idxlist, accountid, user, pwd):
    db = None
    
    try:
        if idxlist is None:
            raise UIException('No selected pin')
        
        if len(idxlist) < 1:
            raise UIException('No selected pin')
        
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        
        params = ','.join(["'{0}'".format(i) for i in idxlist])
        q = 'delete from pinuserid where custid = ? and pin in ({0})'.format(params)
        db.cur.execute(q, accountid)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def pinuserid_deleteall(accountid, user, pwd):
    db = None
    
    try:
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        
        q = 'delete from pinuserid where custid = ?'
        db.cur.execute(q, accountid)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def pinuserid_add(o, user, pwd):
    db = None
    
    try:
        assert isinstance(o, models.PinUserID)
        
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        
        pinuserid_create(db, o)
        db.commit()
            
    finally:
        if db is not None:
            db.dispose()

def pinuserid_update(pin, o, user, pwd):
    db = None
    
    try:
        assert isinstance(o, models.PinUserID)
        
        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        
        q = """
            update pinuserid set pin = ?, description = ?, creditlimit = ? 
            where custid = ? and pin = ? and userid = ?
            """
        db.cur.execute(q, o.pin, o.description, o.creditlimit, o.custid, pin, o.userid)
        db.commit()
        
    except pyodbc.IntegrityError as e:
        if utils.isduplicatekey(e):
            raise UIException('Pin {0} already exist'.format(o.pin))
        
        else:
            raise UIException(traceback.format_exc())
        
    finally:
        if db is not None:
            db.dispose()
            
def rtsms_get(masterid, dbx=None):
    dic = {}
    o = models.RTSMSMaster()
    u = models.RTSMSUser()
    db = dbx
    
    try:
        if db is None:
            db = utils.connectRTSMSDB()
        
        q = """
            select a.masterid, a.enterpriseid, a.wholesalerid, a.rtaccount, a.rtaccounttype, 
            a.smscredit, a.status, a.ip, a.accounttype, a.createdate, 
            a.maxuser, a.maxcontact_admin, a.maxcontact_user, a.maxalert_admin, a.maxalert_user, 
            a.monthlyfee, a.senderid, 
            b.userid, b.usermobile, b.useremail, b.username, b.loginid 
            from rtsms_master a inner join rtsms_user b on a.masterid = b.masterid 
            where b.userlevel = 100 and a.masterid = ?
            """
        r = db.cur.execute(q, masterid).fetchone()
        if r is not None:
            o.masterid = r.masterid
            o.enterpriseid = r.enterpriseid
            o.wholesalerid = r.wholesalerid
            o.rtaccount = r.rtaccount
            o.rtaccounttype = r.rtaccounttype
            o.smscredit = r.smscredit
            o.status = r.status
            o.ip = r.ip
            o.accounttype = r.accounttype
            o.createdate = r.createdate
            o.maxuser = r.maxuser
            o.maxcontact_admin = r.maxcontact_admin
            o.maxcontact_user = r.maxcontact_user
            o.maxalert_admin = r.maxalert_admin
            o.maxalert_user = r.maxalert_user
            o.monthlyfee = r.monthlyfee
            o.senderid = r.senderid
            
            u.userid = r.userid
            u.usermobile = r.usermobile
            u.useremail = r.useremail
            u.username = r.username
            u.loginid = r.loginid
            
            dic['rtsms_master'] = o
            dic['rtsms_user'] = u
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return dic

def rtsms_create(accountid, wholesalerkey, m, u):
    smsmasterid = 0
    b = False
    db = None
    
    try:
        assert isinstance(m, models.RTSMSMaster)
        assert isinstance(u, models.RTSMSUser)
        
        doc = minidom.Document()
        root = doc.createElement('RTSMS_sp_Master_CreateNew')
        root.setAttribute('enterprise_id', str(m.enterpriseid))
        root.setAttribute('wholesaler_id', str(wholesalerkey))
        root.setAttribute('rt_account', str(accountid))
        root.setAttribute('rt_account_type', '1')
        root.setAttribute('initial_credit', '0')
        root.setAttribute('status', '1')
        root.setAttribute('account_type', '1')
        root.setAttribute('admin_name', str(u.username))
        root.setAttribute('admin_email', str(u.useremail))
        root.setAttribute('admin_mobile', str(u.usermobile))
        root.setAttribute('admin_login', str(u.loginid))
        root.setAttribute('admin_password', str(u.loginpwd))
        root.setAttribute('sender_id', str(m.senderid))
        root.setAttribute('max_user', str(m.maxuser))
        root.setAttribute('max_admin_contact', str(m.maxcontact_admin))
        root.setAttribute('max_admin_alert', str(m.maxalert_admin))
        root.setAttribute('max_user_contact', str(m.maxcontact_user))
        root.setAttribute('max_user_alert', str(m.maxalert_user))
        root.setAttribute('monthly_fee', '0')
        
        txt = doc.createTextNode('')
        root.appendChild(txt)
        
        doc.appendChild(root)
        x = doc.toprettyxml(encoding='UTF-16').strip()
        
        db = utils.connectRTSMSDB()
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int,
              @ErrorCode int
            exec @return_value = RTSMS_sp_Master_CreateNew
              @xml = ?,
              @ErrorCode = @ErrorCode output
            select
              @ErrorCode as N'ErrorCode',
              @return_value as N'Ret'
            """
        db.cur.execute(q, x)
        r = db.cur.fetchone()
        n = db.cur.nextset()
         
        if n:
            rx = db.cur.fetchone()
            b, smsmasterid = rtsms_create_check_error(0, rx.MasterID)
            
        else:
            b, smsmasterid = rtsms_create_check_error(r.ErrorCode, r.Ret)
            
    finally:
        if db is not None:
            db.dispose()
            
    return b, smsmasterid

def rtsms_userdetail_update(smsmasterid, accountid):
    db = None
    
    try:
        db = initdb(dbx=None, autocommit=False)
        
        q = """
            Update userdetailext set smsmasterid = ? 
            where accountid = ?
            """
        db.cur.execute(q, smsmasterid, accountid)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def rtsms_update(m, u, chg):
    ls = []
    qs = None
    b = False
    db = None
    
    try:
        assert isinstance(m, models.RTSMSMaster)
        assert isinstance(u, models.RTSMSUser)
        
        doc = minidom.Document()
        root = doc.createElement('RTSMS_sp_Master_Update')
        root.setAttribute('master_id', str(m.masterid))
        root.setAttribute('admin_id', str(u.userid))
        
        if 'admin_name' in chg:
            root.setAttribute('admin_name', str(u.username))
            ls.append('{0}:{1}'.format(chg['admin_name'], str(u.username)))
            
        if 'admin_mobile' in chg:
            root.setAttribute('admin_mobile', str(u.usermobile))
            ls.append('{0}:{1}'.format(chg['admin_mobile'], str(u.usermobile)))
            
        if 'admin_email' in chg:
            root.setAttribute('admin_email', str(u.useremail))
            ls.append('{0}:{1}'.format(chg['admin_email'], str(u.useremail)))
            
        if 'admin_login' in chg:
            root.setAttribute('admin_login', str(u.loginid))
            ls.append('{0}:{1}'.format(chg['admin_login'], str(u.loginid)))
            
        if 'sender_id' in chg:
            root.setAttribute('sender_id', str(m.senderid))
            ls.append('{0}:{1}'.format(chg['sender_id'], str(m.senderid)))
            
        if 'max_user' in chg:
            root.setAttribute('max_user', str(m.maxuser))
            ls.append('{0}:{1}'.format(chg['max_user'], str(m.maxuser)))
            
        if 'max_admin_contact' in chg:
            root.setAttribute('max_admin_contact', str(m.maxcontact_admin))
            ls.append('{0}:{1}'.format(chg['max_admin_contact'], str(m.maxcontact_admin)))
            
        if 'max_admin_alert' in chg:
            root.setAttribute('max_admin_alert', str(m.maxalert_admin))
            ls.append('{0}:{1}'.format(chg['max_admin_alert'], str(m.maxalert_admin)))
            
        if 'max_user_contact' in chg:
            root.setAttribute('max_user_contact', str(m.maxcontact_user))
            ls.append('{0}:{1}'.format(chg['max_user_contact'], str(m.maxcontact_user)))
            
        if 'max_user_alert' in chg:
            root.setAttribute('max_user_alert', str(m.maxalert_user))
            ls.append('{0}:{1}'.format(chg['max_user_alert'], str(m.maxalert_user)))
            
        txt = doc.createTextNode('')
        root.appendChild(txt)
        
        doc.appendChild(root)
        x = doc.toprettyxml(encoding='UTF-16').strip()
        
        db = utils.connectRTSMSDB()
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int,
              @ErrorCode int
            exec @return_value = RTSMS_sp_Master_Update
              @xml = ?,
              @ErrorCode = @ErrorCode output
            select
              @ErrorCode as N'ErrorCode',
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, x).fetchone()
        if r.ErrorCode == 0:
            b = True
            qs = ', '.join(ls)

    finally:
        if db is not None:
            db.dispose()
            
    return b, qs

def rtsms_create_check_error(errcode, ret):
    smsmasterid = 0
    b = False
    
    if errcode != 0:
        if errcode == 2:
            raise UIException('WARNING: Login ID too short, please re-enter the Admin Login with more than 6 characters!')
        
        elif errcode == 3:
            raise UIException('WARNING: Login Password too short, please re-enter the Password with more than 6 characters!')
        
        elif errcode == 5:
            raise UIException('WARNING: Admin Login already in use in SMS account, please re-enter the Admin Login!')
        
        elif errcode == 6:
            raise UIException('WARNING: Admin Login already in use in iFax account, please re-enter the Admin Login!')
        
        else:
            raise Exception('Error: {0}'.format(errcode))
    
    else:
        smsmasterid = ret
        if smsmasterid == 0:
            raise Exception('Error while creating SMS Account. SMS MasterID=0. Pls contact administrator')
        
        else:
            b = True
            
    return b, smsmasterid



def insert_into_sql_log(qs, accountid, username, db, status=0): 
    q = """
        insert into gwlog (status, transactionsql, username, creationdate, accountid) 
        values (?, ?, ?, ?, ?)
        """
    db.cur.execute(q, status, qs, username, datetime.now(), accountid)
            
def topuplog_list(accountid, dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = """
            select indexkey, accountid, topupdate, topupdatetime, topupvalue, oldcreditlimit, newcreditlimit,
            topupby, paylater, topuprequestkey 
            from ttopuplog where accountid = ?
            order by topupdatetime desc
            """
        rows = db.cur.execute(q, accountid).fetchall()
        for r in rows:
            o = models.TopupLog()
            o.indexkey = r.indexkey
            o.accountid = r.accountid
            o.topupdate = r.topupdate
            o.topupdatetime = r.topupdatetime
            o.topupvalue = r.topupvalue
            o.oldcreditlimit = r.oldcreditlimit
            o.newcreditlimit = r.newcreditlimit
            o.topupby = r.topupby
            o.paylater = r.paylater
            o.topuprequestkey = r.topuprequestkey
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l
            
def topuprequest_list(accountid, dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = """
            select a.indexkey, a.requestdate, a.accountid, a.amount, a.balance, a.creator, a.notes, a.posted, a.topuptype as topuptype1,
            b.topuptype, b.sdesc 
            from topuprequest a, topuptype b 
            where a.accountid = ? and a.topuptype = b.topuptype
            order by a.requestdate desc
            """
        rows = db.cur.execute(q, accountid).fetchall()
        for r in rows:
            o = models.TopupRequest()
            o.indexkey = r.indexkey
            o.requestdate = r.requestdate
            o.accountid = r.accountid
            o.amount = r.amount
            o.balance = r.balance
            o.creator = r.creator
            o.notes = r.notes
            o.posted = r.posted
            o.topuptype = r.topuptype1
            
            x = models.TopupType()
            x.topuptypeid = r.topuptype
            x.sdesc = r.sdesc
            
            o.topuptyperef = x
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def postnonepaymenttopup(o):
    db = None
    posted = 0
    ex = None
    lqs = None
    indexkey = None
        
    try:
        assert isinstance(o, models.Topup)
        
        db = initdb(dbx=None, autocommit=False)
        dic = postnonepaymenttopup_check(o, db)
        amt = dic['amt']
        
        if o.regkey == '':
            o.regkey = topuprequest_create(o, amt, db)
                
        else:
            q = 'select posted from topuprequest where indexkey = ?'
            r = db.cur.execute(q, o.regkey)
            if r is not None:
                posted = r.posted
                
            if posted != 0:
                raise UIException('Unable to update this topup request because it just has been posted')
            
        if o.isnonepaymenttopup == True:
            postedtime, ex, lqs = insert_topup_log(db, o, amt, '', dic['paylater'])
            if postedtime is not None:
                # create payment record
                q = """
                    insert into paymentdetail (collectionagency, mode, amount, balance, paymentdate, notes, location, 
                    referenceid, creator, posted) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """
                db.cur.execute(q, dic['colagency'], dic['colmode'], amt, 0, datetime.now(), o.notes, '', o.accountid, o.username, 1)
                r = db.cur.execute("select IDENT_CURRENT('paymentdetail') as 'indexkey'").fetchone()
                if r is not None:
                    indexkey = r.indexkey
                    
                # insert PaymentKnockingDetail record
                q = """
                    insert into paymentknockingdetail (paymentkey, topuprequestkey, amount, posteddate) 
                    values (?, ?, ?, ?)
                    """
                db.cur.execute(q, indexkey, o.regkey, amt, postedtime)
                
                q = """
                    update topuprequest set posted = 1, amount = ?, balance = 0, topuptype = ? 
                    where indexkey = ?
                    """
                db.cur.execute(q, amt, o.topuptype, o.regkey)
                db.commit()
                
        if o.authority != '':
            insert_authority_record(o, db)
            q = """
                update topuprequest set notes = ? where indexkey = ? and notes = ''
                """
            db.cur.execute(q, o.authorityremark, o.regkey)
            db.commit()
                
    finally:
        if db is not None:
            db.dispose()
            
    return ex, lqs
        
def postnonepaymenttopup_check(o, dbx=None):
    assert isinstance(o, models.Topup)
    
    colmode = ''
    paylater = 0
    amt = o.amount
    colagent = o.colagency
    
    if o.topuptype == constants.PAYLATER:
        paylater = 1
        if o.amount <= 0:
            raise UIException('Invalid Pay Later amount, must be greater than 0')
        
        colagent = getpaylater_wholesaler(o.wholesalerkey, dbx)
        if colagent == '':
            raise UIException('This customer does not have Pay Later priviledge')
        
        colmode = 'PAYLATER'
    
    elif o.topuptype == constants.CREDIT_TRANSFER:
        raise UIException('You cannot select Credit Transfer topup type')
    
    elif o.topuptype == constants.TOPDOWN:
        amt = 0 - abs(o.amount)
        colmode = 'AIRTIME'
    
    elif o.topuptype == constants.TERMINATION_REFUND:
        amt = 0 - abs(o.amount)
        colmode = 'AIRTIME'
    
    elif o.topuptype == constants.FREE_AIRTIME:
        colmode = 'AIRTIME'
    
    elif o.topuptype == constants.REFUND:
        colmode = 'AIRTIME'
        
    return {
        'colmode': colmode,
        'paylater': paylater,
        'amt': amt,
        'colagency': colagent
    }
    
def authority_submit(o):
    db = None
    
    try:
        assert isinstance(o, models.Authority)
        
        db = initdb()
        
        q = """
            select sloginpassword from tlogin 
            where sloginname = ?
            """
        r = db.cur.execute(q, o.supervisor).fetchone()
        if r is not None:
            psw = utils.simpledecrypt(r.sloginpassword, 8).upper()
            if o.psw != psw:
                raise UIException('Invalid password')
        
        else:
            raise UIException('Unable to find supervisor: {0}'.format(o.supervisor))
        
    finally:
        if db is not None:
            db.dispose()
    
def topuprequest_create(o, amount, dbx=None):
    db = dbx
    indexkey = None
    
    try:
        assert isinstance(o, models.Topup)
        
        db = initdb(dbx=dbx, autocommit=False)
        
        q = """
            insert into topuprequest (requestdate, accountid, amount, balance, creator, notes, posted, topuptype) 
            values (?, ?, ?, ?, ?, ?, ?, ?)
            """
        db.cur.execute(q, datetime.now(), o.accountid, amount, amount, o.username, o.notes, 0, o.topuptype)
        r = db.cur.execute("select IDENT_CURRENT('topuprequest') as 'indexkey'").fetchone()
        db.commit()
        if r is not None:
            indexkey = r.indexkey
            
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return indexkey

def topuprequest_amount_update(o):
    db = None
    
    try:
        assert isinstance(o, models.Topup)
        
        db = initdb(dbx=None, autocommit=False)
        
        q = """
            update topuprequest set amount = ?, balance = ? 
            where indexkey = ?
            """
        db.cur.execute(q, o.amount, o.amount, o.regkey)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def topuprequest_notes_update(o):
    db = None
    
    try:
        assert isinstance(o, models.Topup)
        
        db = initdb(dbx=None, autocommit=False)
        
        q = """
            update topuprequest set notes = ? 
            where indexkey = ?
            """
        db.cur.execute(q, o.notes, o.regkey)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def topuprequest_topuptype_update(o):
    db = None
    
    try:
        assert isinstance(o, models.Topup)
        
        db = initdb(dbx=None, autocommit=False)
        
        q = """
            update topuprequest set balance = amount, topuptype = ? 
            where indexkey = ?
            """
        db.cur.execute(q, o.topuptype, o.regkey)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def ispostedtopuprequest(regkey):
    b = False
    db = None
    
    try:
        db = initdb()
        
        q = 'select posted from topuprequest where indexkey = ?'
        r = db.cur.execute(q, regkey).fetchone()
        if r is not None:
            k = r.posted
            b = True if k != 0 else False
            
    finally:
        if db is not None:
            db.dispose()
            
    return b

def allowperformtopup(topuptype=1, aclevel=2):
    b = False
    
    if topuptype == 1:
        if aclevel != 2 and aclevel != 4:
            raise UIException('You are not allowd to perform Topup Type request')
        
    b = True
    return b

def getpaylater_wholesaler(wskey, dbx=None):
    s = ''
    db = dbx
    
    try:
        db = initdb(dbx=dbx, autocommit=True)
        
        q = """
            select agentname from collectionagency 
            where paymentmode = 'PAYLATER' and wsid = ?
            """
        r = db.cur.execute(q, wskey).fetchone()
        if r is not None:
            s = r.agentname
            
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return s

def insert_authority_record(o, db):
    assert isinstance(o, models.Topup)
    q = """
        insert into authority (requestkey, supervisor, remark) 
        values (?, ? ?)
        """
    db.cur.execute(q, o.regkey, o.authority, o.authorityremark)
    db.commit()

def insert_topup_log(db, o, amount, postedtime='', paylater=0):
    ex = None
    lqs = None
    dt = None
    
    assert isinstance(o, models.Topup)
        
    if postedtime == '':
        dt = datetime.now()
        postedtime = dt.strftime('%Y-%m-%d %H:%M:%S')
        
    q = """
        select accountid, creditlimit, creditusage, isubtype, igatetype from userdetail 
        where accountid = ?
        """
    r = db.cur.execute(q, o.accountid).fetchone()
    if r.igatetype == constants.GWID.GWALL:
        raise UIException('Invalid GateType :0')
    
    creditlimit = r.creditlimit + amount
    
    q = 'update userdetail set creditlimit = ? where accountid = ?'
    db.cur.execute(q, creditlimit, o.accountid)
    
    q = """
        insert into ttopuplog (accountid, topupdatetime, topupvalue, oldcreditlimit, 
        newcreditlimit, topupby, paylater, topuprequestkey) 
        values (?, ?, ?, ?, ?, ?, ?, ?)
        """
    db.cur.execute(q, o.accountid, dt, amount, r.creditlimit, creditlimit, o.username, paylater, o.regkey)
    
    q = 'update topuprequest set posted = 1, balance = 0 where indexkey = ?'
    db.cur.execute(q, o.regkey)
    
    db.commit()
    
    if o.batchupload == False:
        ex, lqs = remotecreditlimit_update(r.igatetype, o.accountid, creditlimit)
            
    return dt, ex, lqs
            
def remotecreditlimit_update(igatetype, accountid, newlimit):
    qs = None
    db = None
    ex = None
    lqs = []
    
    try:
        q = 'select accountid, creditlimit, creditusage from userdetail where accountid = ? and igatetype = ?'
        prm = [accountid, igatetype]
        db = getremotedata(igatetype, q, prm)
        r = db.cur.fetchone()
        
        if r is not None:
            qs = creditlimit_update(db, igatetype, accountid, r.creditlimit, newlimit)
            
            if qs is None:
                qs = 'Failed to update remote Credit Limit! GatewayID:{0}, AccountID:{1}, NewCreditLimit:{2}'.format(igatetype, accountid, newlimit)
                lqs.append(qs)
                ex = 'Failed to update credit limit at gateway:{0}, AccountID:{1}, amount:{2}, topup will be performed in batch mode, please upload the topup commands manually'.format(igatetype, accountid, newlimit)
                
            else:
                lqs.append(qs)
    
    finally:
        if db is not None:
            db.dispose()
            
    return ex, lqs
            
def creditlimit_update(db, igatetype, accountid, oldlimit, newlimit):
    qs = None
    
    try:
        q = 'update userdetail set creditlimit = ? where accountid = ? and igatetype = ?'
        db.cur.execute(q, newlimit, accountid, newlimit)
        qs = 'UpdateRemoteCreditLimit Gw({0}):AccountID:{1}, Old CreditLimit:{2:.2f}, New Credit Limit:{3:.2f}'.format(igatetype, accountid, oldlimit, newlimit)
        
    except:
        raise
    
    return qs
            
def getremotedata(igatetype, q, prm):
    db = None
    
    try:
        db, gw = utils.connectGWDB(igatetype)
        try:
            db.cur.execute(q, prm)
            
        except:
            raise UIException('Failed to query remote database {0}'.format(gw))
        
    except:
        raise
        
    return db

def supervisor_list(dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = """
            select sloginname, ip from tlogin 
            where iaccesslevel = 2 
            order by sloginname
            """
        rows = db.cur.execute(q).fetchall()
        for r in rows:
            o = models.TLogin()
            o.sloginname = r.sloginname
            o.ip = r.ip
            l.append(o)
            
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def list_topuptype(dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = """
            select topuptype, sdesc, requirepayment from topuptype 
            where selectable = 1 
            order by topuptype
            """
        rows = db.cur.execute(q).fetchall()
        for r in rows:
            o = models.TopupType()
            o.topuptypeid = r.topuptype
            o.sdesc = utils.decode(r.sdesc)
            o.requirepayment = r.requirepayment
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def topuprequest_delete(idxlist):
    db = None
    
    try:
        if idxlist is None:
            raise UIException('No selected topup request')
        
        if len(idxlist) < 1:
            raise UIException('No selected topup request')
        
        db = initdb(dbx=None, autocommit=False)
        
        params = ','.join([str(i) for i in idxlist])
        q = 'delete from topuprequest where posted = 0 and indexkey in ({0})'.format(params)
        db.cur.execute(q)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()
            
def userdetailcontactlist_list(accountid, dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = """
            select idx, accountid, sname, ipersonincharge, srace, sposition, sphone, sfax, smobile, semail
            from userdetailcontactlist
            where accountid = ?
            order by sname
            """
        rows = db.cur.execute(q, accountid).fetchall()
        for r in rows:
            o = models.UserDetailContactList()
            o.idx = r.idx
            o.accountid = r.accountid
            o.sname = r.sname
            o.ipersonincharge = r.ipersonincharge
            o.srace = r.srace
            o.sposition = r.sposition
            o.sphone = r.sphone
            o.sfax = r.sfax
            o.smobile = r.smobile
            o.semail = r.semail
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def userdetailcontactlist_create(o):
    db = None
    
    try:
        assert isinstance(o, models.UserDetailContactList)
        
        db = initdb(dbx=None, autocommit=False)
        
        q = """
            insert into userdetailcontactlist (accountid, sname, ipersonincharge, srace, sposition, 
            sphone, sfax, smobile, semail) 
            values (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
        db.cur.execute(q, o.accountid, o.sname, o.ipersonincharge, o.srace, o.sposition, 
                       o.sphone, o.sfax, o.smobile, o.semail)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def userdetailcontactlist_update(o):
    db = None
    
    try:
        assert isinstance(o, models.UserDetailContactList)
        
        db = initdb(dbx=None, autocommit=False)
        
        q = """
            update userdetailcontactlist set sname = ?, ipersonincharge = ?, srace = ?, 
            sposition = ?, sphone = ?, sfax = ?, smobile = ?, semail = ? 
            where idx = ? and accountid = ?
            """
        db.cur.execute(q, o.sname, o.ipersonincharge, o.srace, o.sposition, o.sphone, o.sfax,
                       o.smobile, o.semail, o.idx, o.accountid)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def userdetailcontactlist_delete(idxlist):
    db = None
    
    try:
        if idxlist is None:
            raise UIException('No selected contact')
        
        if len(idxlist) < 1:
            raise UIException('No selected contact')
        
        db = initdb(dbx=None, autocommit=False)
        
        params = ','.join([str(i) for i in idxlist])
        q = 'delete from userdetailcontactlist where idx in ({0})'.format(params)
        db.cur.execute(q)
        db.commit()
    
    finally:
        if db is not None:
            db.dispose()
            
def load_type(maintype):
    dic = {}
    db = None
    
    try:
        db = initdb()
        
        dic['maintype'] = list_maintype(db)
        dic['gatetype'] = list_gatetype(db)
        dic['subtype'] = list_subtype(maintype, db)
        dic['ratetype'] = list_ratetype(maintype, db)
        dic['lcrtype'] = list_lcrtype(db)
    
    finally:
        if db is not None:
            db.dispose()
    
    return dic

def load_techinfolookup():
    dic = {}
    db = None
    
    try:
        db = initdb()
        
        dic['supportteam'] = list_supportteamlist(db)
    
    finally:
        if db is not None:
            db.dispose()
            
    return dic

def load_gatewaylookup(wskey='0'):
    dic = {}
    db = None
    
    try:
        db = initdb()
        
        dic['ws'] = list_wholesaler(db, wskey)
        dic['sipsubtype'] = list_sipsubtype(db)
        dic['siplcrtype'] = list_siplcrtype(db)
    
    finally:
        if db is not None:
            db.dispose()
            
    return dic

def load_codearea():
    dic = {}
    db = None
    
    try:
        db = initdb()
        
        dic['ca'] = list_area(db)
        # dic['sipsubtype'] = list_sipsubtype(db)
        # dic['siplcrtype'] = list_siplcrtype(db)
    
    finally:
        if db is not None:
            db.dispose()
            
    return dic


def list_siplcrtype(dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = """
            select siplcrtype, sdesc from tsiplcrtype
            order by siplcrtype
            """
        rows = db.cur.execute(q).fetchall()
        for r in rows:
            o = models.SipLCRType()
            o.siplcrtype = r.siplcrtype
            o.sdesc = utils.decode(r.sdesc)
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def list_sipsubtype(dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = """
            select sipsubtype, sdesc from tsipsubtype
            order by sipsubtype
            """
        rows = db.cur.execute(q).fetchall()
        for r in rows:
            o = models.SipSubType()
            o.sipsubtype = r.sipsubtype
            o.sdesc = utils.decode(r.sdesc)
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def list_devicelist(dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = """
            select deviceid, devicename from devicelist
            order by deviceid
            """
        rows = db.cur.execute(q).fetchall()
        for r in rows:
            o = models.DeviceList()
            o.deviceid = r.deviceid
            o.devicename = r.devicename
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def list_supportteamlist(dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = """
            select teamid, teamname from supportteamlist
            order by teamid
            """
        rows = db.cur.execute(q).fetchall()
        for r in rows:
            o = models.SupportTeamList()
            o.teamid = r.teamid
            o.teamname = r.teamname
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def list_wholesaler(dbx=None, wskey='0'):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = """
            select wholesalerkey, wholesalername from twholesaler
            """
        lq = [q]
        if wskey not in constants.EMPTY_WHOLESALER:
            lq.append(' where wholesalerkey = ?')

        lq.append(' order by wholesalerkey')
        q = ''.join(lq)

        if wskey not in constants.EMPTY_WHOLESALER:
            db.cur.execute(q, wskey)

        else:
            db.cur.execute(q)

        rows = db.cur.fetchall()
        for r in rows:
            o = models.Wholesaler()
            o.wholesalerkey = r.wholesalerkey
            o.wholesalername = utils.decode(r.wholesalername)
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l 

def list_lcrtype(dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = "select ilcrtype, sname from tlcrtype order by ilcrtype desc"
        rows = db.cur.execute(q).fetchall()
        for r in rows:
            o = models.LCRType()
            o.ilcrtype = r.ilcrtype
            o.sname = utils.decode(r.sname)
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def list_ratetype(maintype, dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
        
        if maintype != 0:
            q = """
                select iratetype, sname from tratetype
                where imaintype = ?
                order by iratetype
                """
            db.cur.execute(q, maintype)
                
        else:
            q = """
                select iratetype, sname from tratetype
                order by iratetype
                """
            db.cur.execute(q)
            
        rows = db.cur.fetchall()
        for r in rows:
            o = models.RateType()
            o.iratetype = r.iratetype
            o.sname = utils.decode(r.sname)
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def list_subtype(maintype, dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
            
        if maintype != 0:
            q = """
                select isubtype, sname, sfeaturescode from tsubtype
                where imaintype = ?
                order by isubtype
                """
            db.cur.execute(q, maintype)
            
        else:
            q = """
                select isubtype, sname, sfeaturescode from tsubtype
                order by isubtype
                """
            db.cur.execute(q)
            
        rows = db.cur.fetchall()
        for r in rows:
            o = models.SubType()
            o.isubtype = r.isubtype
            o.sname = utils.decode(r.sname)
            o.sfeaturescode = r.sfeaturescode
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def list_gatetype(dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = """
            select igatetype, sname from tgatetype
            order by igatetype
            """
        rows = db.cur.execute(q).fetchall()
        for r in rows:
            o = models.GateType()
            o.igatetype = r.igatetype
            o.sname = utils.decode(r.sname)
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def list_maintype(dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = "select imaintype, sname from tmaintype order by imaintype"
        rows = db.cur.execute(q).fetchall()
        for r in rows:
            o = models.MainType()
            o.imaintype = r.imaintype
            o.sname = utils.decode(r.sname)
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def userdetail_load(accountid,gLogin_Wholesaler):
    dic = {}
    db = None
    
    try:
        db = initdb()
        dic['userdetail'] = userdetail_get(accountid, gLogin_Wholesaler, db)
        dic['availablenum'] = userdetail_count_available(accountid, db)
        dic['usednum'] = userdetail_count_used(accountid, db)
        dic['suspendnum'] = userdetail_count_suspend(accountid, db)

        # accountid='999990205'

        # dic['userdetail'] = userdetail_getbatch(batch_id, db)
        # dic['callerid'] = authentication_list(accountid, db)
        # dic['batchid'] = batch_list(batch_id, db) 

        dic['rt015num'] = bbauthentication_list_available(accountid, db)
        dic['rt015used'] = bbauthentication_list_used(accountid, db)
        dic['rt015block'] = bbauthentication_list_blocked(accountid, db)
        dic['rt015test'] = bbauthentication_list_testing(accountid, db)

        # dic['pin'] = pinuserid_list(accountid, db)
        # dic['topuprequest'] = topuprequest_list(accountid, db)
        # dic['topuphistory'] = topuplog_list(accountid, db)
        # dic['remark'] = userdetailremark_list(accountid, db)
        # dic['contactlist'] = userdetailcontactlist_list(accountid, db)
        # dic['devicelist'] = vuserdetaildevice_list(accountid, db)
    
    finally:
        if db is not None:
            db.dispose()
            
    return dic

# def userdetail_list(i, keyword, dbx=None):
#     l = []
#     db = dbx
    
#     try:
#         db = initdb(dbx)
        
#         q = ''
#         prm = ['%{0}%'.format(keyword)]
#         print(prm)
        
                
#         if i == 1: #phone
#             print("ok callerid")
#             q = """
#                 select bn.batch_id, cid.batchid, bn.batch_date,bn.batch_qty, bn.uploadby,cid.callerid, cid.assigndate, cid.blockdate, cid.releasedate, 
#                 cid.wskey, cid.status, cid.code_area, cid.state  from gw_BatchNumber bn, GW_callerid cid
#                 where bn.batch_id=cid.batchid and cid.callerid like ?
#                 """
#             lq = [q]


#             lq.append('order by bn.batch_date')
#             q = ''.join(lq)
            
                
#         else:
#             print("not callerid")

#             q = """
#                 select top 10000 batch_id, batch_date, batch_qty, uploadby from gw_BatchNumber 
#                 where batch_id like ?
#                 """
#             lq = [q]


#             lq.append('order by batch_date')
#             q = ''.join(lq)


#         params = tuple(prm)
                
#         rows = db.cur.execute(q, params).fetchall()
#         for r in rows:
#             o = models.batchID()
#             o.batch_id = r.batch_id
#             o.batch_qty = r.batch_qty
#             o.batch_date = r.batch_date

            
#             if i == 1:
                
#                 o = models.callerid_detail()
#                 o.callerid = r.callerid
#                 o.batchid = r.batch_id
#                 print(o.batchid)

                
#             l.append(o)
    
#     finally:
#         if db is not None and dbx is None:
#             db.dispose()
            
#     return l

# def userdetail_list(i, keyword, dbx=None):
#     l = []
#     db = dbx
    
#     try:
#         db = initdb(dbx)
        
#         q = ''
#         prm = ['%{0}%'.format(keyword)]
        
#         if i == 0: #custid
#             q = """
#                 select top 10000 batch_id, batch_date, batch_qty, uploadby from gw_BatchNumber 
#                 where batch_id like ?
#                 """
#             lq = [q]

            
#             lq.append(' order by batch_date')
#             q = ''.join(lq)
                
#         elif i == 1: #phone
#             print("here")
#             q = """
#                 select top 10000 a.callerid,  a.batchid from gw_callerid a,
#                 gw_BatchNumber b where a.batchid = b.batch_id and a.callerid like ?               
#                 """
#             lq = [q]


#             lq.append(' order by a.callerid')
#             q = ''.join(lq)
#             print(q)
                
#         else:
#             print("here2")

#             q = """
#                 select top 10000 batch_id, batch_date, batch_qty, uploadby from gw_BatchNumber 
#                 where batch_id like ?
#                 """
#             lq = [q]

            
#             lq.append(' order by batch_date')
#             q = ''.join(lq)

#         params = tuple(prm)
                
#         rows = db.cur.execute(q, params).fetchall()
#         for r in rows:
#             if i == 1:
            
#                 o = models.callerid_detail()
#                 o.callerid = r.callerid
#                 o.batchid = r.batchid

 
#             else:
        
#                 o = models.batchID()
#                 o.batch_id = r.batch_id
#                 o.batch_qty = r.batch_qty
#                 o.batch_date = r.batch_date
            
               
#                 l.append(o)
    
#     finally:
#         if db is not None and dbx is None:
#             db.dispose()
            
#     return l

def userdetail_list(i, keyword, dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = ''
        prm = ['%{0}%'.format(keyword)]
        print(prm)
        
                
        if i == 1: #phone
            print("ok callerid")
            q = """
                select bn.batch_id, cid.batchid, bn.batch_date,bn.batch_qty, bn.uploadby,cid.callerid, cid.assigndate, cid.blockdate, cid.releasedate, 
                cid.wskey, cid.status, cid.code_area, cid.state  from gw_BatchNumber bn, GW_callerid cid
                where bn.batch_id=cid.batchid and cid.callerid like ?
                """
            lq = [q]


            lq.append('order by bn.batch_date')
            q = ''.join(lq)
            
                
        else:
            print("not callerid")

            q = """
                select top 10000 batch_id, batch_date, batch_qty, uploadby from gw_BatchNumber 
                where batch_id like ?
                """
            lq = [q]


            lq.append('order by batch_date')
            q = ''.join(lq)


        params = tuple(prm)
                
        rows = db.cur.execute(q, params).fetchall()
        for r in rows:
            # o = models.UserDetail()
            # o.accountid = r.accountid
            # o.name = utils.decode(r.name)batchID
            o = models.batchID()
            o.batch_id = r.batch_id
            o.batch_qty = r.batch_qty
            o.batch_date = r.batch_date
            #o.name = utils.decode(r.name)batchID

            
            if i == 1:
                
                o = models.callerid_detail()
                o.callerid = r.callerid
                o.batchid = r.batch_id
                print(o.batchid)

                
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l


def userdetail_list_by_accountid(accountid, wskey='0'):
    l = []
    db = None

    try:
        db = initdb()

        if len(accountid) > 9:
            accountid = accountid[:9]

        q = """
            select top {0} accountid, name from userdetail 
            where accountid like ?
            """.format(constants.LIST_SIZE)
        lq = [q]
        prm = ['{0}%'.format(accountid)]

        if wskey not in constants.EMPTY_WHOLESALER:
            lq.append(' and accountid in (select accountid from userdetailext where wholesalerkey = ?)')
            prm.append(wskey)

        lq.append(' order by accountid')
        q = ''.join(lq)

        params = tuple(prm)

        rows = db.cur.execute(q, params).fetchall()
        for r in rows:
            l.append(r.accountid)

    finally:
        if db is not None:
            db.dispose()

    return l

def userdetail_get(batch_id, gLogin_Wholesaler, dbx=None):
    o = models.BatchDetail()
    db = dbx
    
    try:
        db = initdb(dbx)
        
      
        q = """
            
	      select batch_id, uploadby, batch_date, batch_qty from gw_batchnumber
            where batch_id like ?
            """
        lq = [q]
        
            
        q = ''.join(lq)
        # r = db.cur.execute(q, batch_id).fetchone()
        r = db.cur.execute(q,'%{0}%'.format(batch_id)).fetchone()

        if r is not None:
            o.batch_id = r.batch_id
            o.uploadby = r.uploadby
            o.uploadby = r.uploadby
            o.batch_date = r.batch_date
            o.batch_qty = r.batch_qty
    
            
        else:
            raise UIException('Record not found')
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return o

def userdetail_count_available(batch_id, dbx=None):
    o = 0
    db = dbx

    try:
        db = initdb(dbx)

        q = """
            select count(cid.status) as availablenum from gw_BatchNumber bn, GW_callerid cid
            where bn.batch_id=cid.batchid and bn.batch_id like ? and  cid.status = 0
            """
        r = db.cur.execute(q, batch_id).fetchone()
        if r is not None:
            o = r.availablenum

    finally:
        if db is not None and dbx is None:
            db.dispose()

    return o

# def callerdetail_get(callerid, dbx=None):
#     o = models.BatchDetail()
#     db = dbx

#     try:
#         db = initdb(dbx)

#         q = """
#             select cid.callerid,cid.batchid, cid.wskey, cid.state, cid,code_area, cid.status from gw_BatchNumber bn, GW_callerid cid
#             where bn.batch_id=cid.batchid and cid.callerid like ? 
#             """
#         r = db.cur.execute(q, callerid).fetchone()
#         if r is not None:
#             o.callerid = r.callerid
#             o.batchid = r.batchid
#             o.wskey = r.wskey
#             o.state = r.state
#             o.status = r.status
#             o.code_area = r.code_area

#             # insert_into_deleted_record(qs, accountid, username, db, -1)


#     finally:
#         if db is not None and dbx is None:
#             db.dispose()

#     return o



def userdetail_count_used(batch_id, dbx=None):
    o = 0
    db = dbx

    try:
        db = initdb(dbx)

        q = """
            select count(cid.status) as usednum from gw_BatchNumber bn, GW_callerid cid
            where bn.batch_id=cid.batchid and bn.batch_id like ? and  cid.status = 1
            """
        r = db.cur.execute(q, batch_id).fetchone()
        if r is not None:
            o = r.usednum

    finally:
        if db is not None and dbx is None:
            db.dispose()

    return o



def userdetail_count_suspend(batch_id, dbx=None):
    o = 0
    db = dbx

    try:
        db = initdb(dbx)

        q = """
            select count(cid.status) as suspendnum from gw_BatchNumber bn, GW_callerid cid
            where bn.batch_id=cid.batchid and bn.batch_id like ? and  cid.status = 3
            """
        r = db.cur.execute(q, batch_id).fetchone()
        if r is not None:
            o = r.suspendnum

    finally:
        if db is not None and dbx is None:
            db.dispose()

    return o




def userdetail_getbatch(batch_id, dbx=None):
    #o = models.UserDetail()
    o = models.BatchDetail()

    db = dbx
    
    try:
        db = initdb(dbx)
    
        q = """
            select batch_id, uploadby, batch_remark,batch_date 
            from gw_batchnumber
            where batch_id= ?
            """
        # print(q)
        r = db.cur.execute(q,'%{0}%'.format(batch_id)).fetchone()
        if r is not None:
            o.batch_id = r.batch_id
            o.uploadby = r.uploadby
            o.batch_remark = r.batch_remark
            o.batch_date = r.batch_date
            print(o.batch_id)
          
           
    finally:
        if db is not None and dbx is None:
            db.dispose()
    # print(o)       
    return o



def userdetail_loadmobusage(accountid):
    v = 0
    db = None
    
    try:
        db = utils.connectAMDB1()
        
        q = """
            SET NOCOUNT ON
            declare 
              @return_value int,
              @Usage money
            exec @return_value = sp_GetIDDUsage_Mobile
              @AccountID = ?,
              @Usage = @Usage output
            select 
              @Usage as N'Usage',
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, accountid).fetchone()
        if r.Ret == 1:
            v = r.Usage
    
    finally:
        if db is not None:
            db.dispose()
            
    return v

def userdetail_loadstdusage(accountid):
    v = 0
    db = None
    
    try:
        db = utils.connectAMDB1()
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int,
              @Usage money
            exec @return_value = sp_GetIDDUsage_STD
              @AccountID = ?,
              @Usage = @Usage output
            select
              @Usage as N'Usage',
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, accountid).fetchone()
        if r.Ret == 1:
            v = r.Usage
    
    finally:
        if db is not None:
            db.dispose()
            
    return v

def userdetail_loadiddusage(accountid):
    v = 0
    db = None
    
    try:
        db = utils.connectAMDB1()
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int,
              @IDDUsage money
            exec @return_value = sp_GetIDDUsage
              @AccountID = ?,
              @IDDUsage = @IDDUsage output
            select
              @IDDUsage as N'IDDUsage',
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, accountid).fetchone()
        if r.Ret == 1:
            v = r.IDDUsage
    
    finally:
        if db is not None:
            db.dispose()
            
    return v

def userdetail_resetmobusage(accountid):
    db = None
    
    try:
        db = utils.connectAMDB1()
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int
            exec @return_value = sp_ResetIDDUsage_Mobile
              @AccountID = ?
            select
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, accountid).fetchone()
        if r.Ret != 1:
            raise Exception('Failed to reset MOB Usage')
    
    finally:
        if db is not None:
            db.dispose()

def userdetail_resetstdusage(accountid):
    db = None
    
    try:
        db = utils.connectAMDB1()
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int
            exec @return_value = sp_ResetIDDUsage_STD
              @AccountID = ?
            select
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, accountid).fetchone()
        if r.Ret != 1:
            raise Exception('Failed to reset STD Usage')
    
    finally:
        if db is not None:
            db.dispose()

def userdetail_resetiddusage(accountid):
    db = None
    
    try:
        db = utils.connectAMDB1()
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int
            exec @return_value = sp_ResetIDDUsage
              @AccountID = ?
            select
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, accountid).fetchone()
        if r.Ret != 1:
            raise Exception('Failed to reset IDD Usage')
    
    finally:
        if db is not None:
            db.dispose()

def userdetailremark_list(accountid, dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = """
            select rmk_index, rmk_accountid, rmk_time, rmk_login, rmk_desc 
            from userdetailremark 
            where rmk_accountid = ?
            order by rmk_time
            """
        rows = db.cur.execute(q, accountid).fetchall()
        for r in rows:
            o = models.UserDetailRemark()
            o.index = r.rmk_index
            o.accountid = r.rmk_accountid
            o.time = r.rmk_time
            o.login = r.rmk_login
            o.desc = r.rmk_desc
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l
            
def userdetailremark_create(o):
    db = None
    
    try:
        assert isinstance(o, models.UserDetailRemark)
        
        db = initdb(dbx=None, autocommit=False)
        
        q = """
            insert into userdetailremark (rmk_accountid, rmk_time, rmk_login, rmk_desc) 
            values (?, ?, ?, ?)
            """
        db.cur.execute(q, o.accountid, datetime.now(), o.login, o.desc)
        db.commit()
    
    finally:
        if db is not None:
            db.dispose()

def vuserdetaildevice_list(accountid, dbx=None):
    l = []
    db = dbx
    
    try:
        db = initdb(dbx)
        
        q = """
            select idx, accountid, deviceid, units, remarks, loginid, devicename
            from v_UserDetailDevice
            where accountid = ?
            order by idx
            """
        rows = db.cur.execute(q, accountid).fetchall()
        for r in rows:
            o = models.vUserDetailDevice()
            o.idx = r.idx
            o.accountid = r.accountid
            o.deviceid = r.deviceid
            o.units = r.units
            o.remarks = r.remarks
            o.loginid = r.loginid
            o.devicename = r.devicename
            l.append(o)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def vuserdetaildevice_create(o):
    db = None
    
    try:
        assert isinstance(o, models.vUserDetailDevice)
        
        db = initdb(dbx=None, autocommit=False)
        
        q = """
            insert into userdetaildevice (accountid, deviceid, units, remarks, loginid) 
            values (?, ?, ?, ?, ?)
            """
        db.cur.execute(q, o.accountid, o.deviceid, o.units, o.remarks, o.loginid)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def vuserdetailldevice_update(o):
    db = None
    
    try:
        assert isinstance(o, models.vUserDetailDevice)
        
        db = initdb(dbx=False, autocommit=False)
        
        q = """
            update userdetaildevice set deviceid = ?, units = ?, remarks = ? 
            where idx = ?
            """
        db.cur.execute(q, o.deviceid, o.units, o.remarks, o.idx)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def vuserdetaildevice_delete(idxlist):
    db = None
    
    try:
        if idxlist is None:
            raise UIException('No selected device')
        
        if len(idxlist) < 1:
            raise UIException('No selected device')
        
        db = initdb(dbx=None, autocommit=False)
        
        params = ','.join([str(i) for i in idxlist])
        q = 'delete from userdetaildevice where idx in ({0})'.format(params)
        db.cur.execute(q)
        db.commit()
    
    finally:
        if db is not None:
            db.dispose()

def wscountries_update(o):
    db = None
    
    try:
        assert isinstance(o, models.WsCountries)
        
        db = initdb()
        
        if o.isvalidcountrycodes == False:
            raise UIException('Country code contain invalid character')
        
        if o.enabled == 9 and o.allow == 0 and o.country == '':
            q = """
                SET NOCOUNT ON
                declare
                  @return_value int,
                  @Msg int
                exec @return_value = sp_WsCountries_Remove
                  @CustID = ?,
                  @Username = ?,
                  @Msg = @Msg output
                select
                  @Msg as N'Msg',
                  @return_value as N'Ret'
                """
            r = db.cur.execute(q, o.custid, o.username).fetchone()
            if r.Ret != 1:
                raise Exception('Failed to update Country code')
        
        else:
            q = """
                SET NOCOUNT ON
                declare
                  @return_value int,
                  @Msg int
                exec @return_value = sp_WsCountries_Add
                  @CustID = ?,
                  @Allow = ?,
                  @Countries = ?,
                  @Username = ?,
                  @Msg = @Msg output
                select
                  @Msg as N'Msg',
                  @return_value as N'Ret'
                """
            r = db.cur.execute(q, o.custid, o.allowtocall, o.country, o.username).fetchone()
            if r.Ret != 1:
                raise Exception('Failed to update Country code')
        
    finally:
        if db is not None:
            db.dispose()

def ifaxuser_list(accountid, wholesalerkey, dbx=None):
    l = []
    db = dbx
    
    try:
        if db is None:
            db = utils.connectAMDB1()
        
        q = """
            SET NOCOUNT ON
            declare 
              @return_value int
            exec @return_value = fx_PostpaidListByAccountId
              @accountId = ?,
              @agentgroup = ?
            select 
              @return_value as N'Ret'
            """
        rows = db.cur.execute(q, accountid, wholesalerkey).fetchall()
        for r in rows:
            o = models.UserDetailFax()
            o.guseridx = r.guseridx
            o.parent_guseridx = r.parent_guseridx
            o.username = r.username
            o.ddi = r.ddi
            o.status = r.status
            o.lastlogin = r.lastlogin
            o.fax_useridx = r.fax_useridx
            l.append(o)
        
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def ifaxuser_get(useridx):
    o = models.FaxDetail()
    db = None
    
    try:
        db = utils.connectAMDB1()
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int,
              @errcode int,
              @errmsg varchar(100)
            exec @return_value = fx_PostpaidUserGetInfo
              @gUseridx = ?,
              @errcode = @errcode output,
              @errmsg = @errmsg output
            select
              @errcode as N'errcode',
              @errmsg as N'errmsg',
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, useridx).fetchone()
            
        b = db.cur.nextset()
        if b:
            rx = db.cur.fetchone()
            if rx.Ret != 1:
                raise Exception('Error: {0}, {1}'.format(rx.errcode, rx.errmsg))
            
            else:
                if r is not None:
                    o.in_enabled = r.in_enabled
                    o.in_notify_num = r.in_notify_num
                    o.in_format = r.in_format
                    o.in_primaryemail = r.in_primaryemail
                    o.out_enabled = r.out_enabled
                    o.out_dailylimit = r.out_dailylimit
                    o.out_notify = r.out_notify
                    o.email = r.email
                    o.status = r.status
                    o.tsi = r.tsi
                    o.custom_header = r.custom_header
                    o.password_retry = r.password_retry
                    o.last_used = r.last_used
                    o.registered = r.registered
                    o.ddi = r.ddi
                    o.fax_useridx = r.fax_useridx
        
    finally:
        if db is not None:
            db.dispose()
            
    return o

def ifax_postpaiduserassignnumber(guseridx, faxnum, wholesalerkey):
    db = None
    
    try:
        db = utils.connectAMDB1()
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int,
              @errorcode int
            exec @return_value = fx_PostpaidUserAssignNumber
              @gUserIdx = ?,
              @agentgroup = ?,
              @faxnum = ?,
              @errorcode = @errorcode output
            select
              @errorcode as N'errorcode',
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, guseridx, wholesalerkey, faxnum).fetchone()
        if r.Ret != 1:
            raise Exception('Error: {0}'.format(r.errorcode))
            
    finally:
        if db is not None:
            db.dispose()

def ifax_postpaiduserunassignnumber(guseridx, faxnum, wholesalerkey):
    db = None
    
    try:
        db = utils.connectAMDB1()
       
        q = """
            SET NOCOUNT ON
            declare
              @return_value int,
              @errorcode int
            exec @return_value = fx_PostpaidUserUnAssignNumber
              @gUserIdx = ?,
              @agentgroup = ?,
              @faxnum = ?,
              @errorcode = @errorcode output
            select
              @errorcode as N'errorcode',
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, guseridx, wholesalerkey, faxnum).fetchone()
        if r.Ret != 1:
            raise Exception('Error: {0}'.format(r.errorcode))
            
    finally:
        if db is not None:
            db.dispose()

def ifaxuser_create(accountid, masterid, username, loginid, password, email):
    l = []
    db = None
    
    try:
        doc = minidom.Document()
        root = doc.createElement('RTSMS_sp_User_New')
        root.setAttribute('master_id', str(masterid))
        root.setAttribute('name', str(username).strip())
        root.setAttribute('login_id', str(loginid).strip())
        root.setAttribute('password', str(password).strip())
        root.setAttribute('email', str(email).strip())
        root.setAttribute('mobile', '')
        root.setAttribute('level', '1')
        root.setAttribute('month_limit', '100')
        root.setAttribute('admin_id', '0')
        
        txt = doc.createTextNode('')
        root.appendChild(txt)
        
        doc.appendChild(root)
        x = doc.toprettyxml(encoding='UTF-8').strip()
        
        db = utils.connectAMDB1()
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int,
              @ErrorMessage varchar(200)
            exec @return_value = RTSMS_sp_Postpaid_NewUser
              @xml = ?,
              @ErrorMessage = @ErrorMessage output
            select
              @ErrorMessage as N'ErrorMessage',
              @return_value as N'Ret'
            """
        db.cur.execute(q, x)
        r = db.cur.fetchone()
        n = db.cur.nextset()
        
        if n:
            rx = db.cur.fetchone()
            if rx.Ret != 1:
                raise Exception('Error: {0}'.format(rx.ErrorMessage))
            
            else:
                l = ifaxuser_list(accountid, 1, db)
                
        else:
            if r.Ret != 1:
                raise Exception('Error: {0}'.format(r.ErrorMessage))
            
            else:
                l = ifaxuser_list(accountid, 1, db)
            
    finally:
        if db is not None:
            db.dispose()
            
    return l

def ifaxuser_delete(guseridx, accountid):
    l = []
    db = None
    
    try:
        db = utils.connectAMDB1()
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int,
              @errcode int,
              @errmsg varchar(200)
            exec @return_value = fx_PostpaidUserDelete
              @gUserIdx = ?,
              @errcode = @errcode output,
              @errmsg = @errmsg output
            select
              @errcode as N'errcode',
              @errmsg as N'errmsg',
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, guseridx).fetchone()
        if r.Ret != 1:
            raise Exception('Error: {0}, {1}'.format(r.errcode, r.errmsg))
        
        else:
            l = ifaxuser_list(accountid, 1, db)
            
    finally:
        if db is not None:
            db.dispose()
    
    return l

def ifaxuser_update(guseridx, o):
    db = None
    
    try:
        assert isinstance(o, models.FaxDetail)
        
        doc = minidom.Document()
        root = doc.createElement('fx_PostpaidUserSetInfo')
        root.setAttribute('in_enabled', str(o.in_enabled))
        root.setAttribute('in_notify_num', str(o.in_notify_num))
        root.setAttribute('in_format', str(o.in_format))
        root.setAttribute('in_primaryemail', str(o.in_primaryemail))
        root.setAttribute('out_enabled', str(o.out_enabled))
        root.setAttribute('out_dailylimit', str(o.out_dailylimit))
        root.setAttribute('out_notify', str(o.out_notify))
        root.setAttribute('email', str(o.email))
        root.setAttribute('status', str(o.status))
        root.setAttribute('tsi', str(o.tsi))
        root.setAttribute('custom_header', str(o.custom_header))
        root.setAttribute('password_retry', str(o.password_retry))
        
        txt = doc.createTextNode('')
        root.appendChild(txt)
        
        doc.appendChild(root)
        x = doc.toprettyxml(encoding='UTF-8').strip()
        
        db = utils.connectAMDB1()
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int,
              @errcode int,
              @errmsg varchar(100)
            exec @return_value = fx_PostpaidUserSetInfo
              @gUserIdx = ?,
              @xmlstr = ?,
              @errcode = @errcode output,
              @errmsg = @errmsg output
            select
              @errcode as N'errcode',
              @errmsg as N'errmsg',
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, guseridx, x).fetchone()
        if r.Ret != 1:
            raise Exception('Error: {0}, {1}'.format(r.errcode, r.errmsg))
        
    finally:
        if db is not None:
            db.dispose()

def postpaidfreenumber_list(prefix, agentgroup):
    l = []
    db = None
    
    try:
        db = utils.connectAMDB1()
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int
            exec @return_value = fx_PostpaidListFreeNumbers
              @prefix = ?,
              @numtype = ?,
              @maxrows = ?,
              @agentgroup = ?,
              @mode = ?
            select
              @return_value as N'Ret'
            """
        rows = db.cur.execute(q, prefix, 0, 200, agentgroup, 1).fetchall()
        for r in rows:
            l.append({ 'faxnum': r.faxnum })
            
    finally:
        if db is not None:
            db.dispose()
            
    return l

def topupreport_list_cust(accountid):
    l = []
    db = None
    
    try:
        db = initdb()
        custid = accountid.upper().replace('X', '_')
        
        q = """
            select accountid, name from userdetail where accountid like ? 
            order by name
            """
        rows = db.cur.execute(q, custid).fetchall()
        for r in rows:
            l.append({
                'accountid': r.accountid,
                'name': utils.decode(r.name)
            })
            
    finally:
        if db is not None:
            db.dispose()
            
    return l

def topupreport_list_cust_by_ws(wholesalerkey):
    l = []
    db = None
    
    try:
        db = initdb()
        
        q = """
            select a.accountid, b.name from userdetailext a, userdetail b 
            where wholesalerkey = ? and b.accountid = a.accountid 
            order by a.accountid
            """
        rows = db.cur.execute(q, wholesalerkey).fetchall()
        for r in rows:
            l.append({
                'accountid': r.accountid,
                'name': utils.decode(r.name)
            })
            
    finally:
        if db is not None:
            db.dispose()
            
    return l

def topupreport_query():
    q = """
        select userdetail.accountid, userdetail.name, ttopuplog.topupdatetime, ttopuplog.topupvalue, 
        ttopuplog.oldcreditlimit, ttopuplog.newcreditlimit, 
        ttopuplog.topupby, ttopuplog.paylater, topuprequesttype.typedesc, topuprequest.notes 
        from userdetail, ttopuplog, topuprequest, topuprequesttype 
        where userdetail.accountid = ttopuplog.accountid 
        and ttopuplog.topuprequestkey = topuprequest.indexkey 
        and topuprequest.topuptype = topuprequesttype.topuptype 
        and ttopuplog.topupdatetime >= ? and ttopuplog.topupdatetime < ?
        order by ttopuplog.accountid, ttopuplog.topupdatetime
        """
    return q

def topupreport_query_selected(idxlist, wholesalerkey=0):
    if wholesalerkey == 0:
        params = ','.join(["'{0}'".format(i) for i in idxlist])
        q = """
            select userdetail.accountid, userdetail.name, ttopuplog.topupdatetime, ttopuplog.topupvalue, 
            ttopuplog.oldcreditlimit, ttopuplog.newcreditlimit, 
            ttopuplog.topupby, ttopuplog.paylater, topuprequesttype.typedesc, topuprequest.notes 
            from userdetail, ttopuplog, topuprequest, topuprequesttype 
            where userdetail.accountid in ({0}) 
            and userdetail.accountid = ttopuplog.accountid 
            and ttopuplog.topuprequestkey = topuprequest.indexkey 
            and topuprequest.topuptype = topuprequesttype.topuptype
            and ttopuplog.topupdatetime >= ? and ttopuplog.topupdatetime < ?
            order by ttopuplog.accountid, ttopuplog.topupdatetime
            """.format(params)
            
    else:
        q = """
            select userdetail.accountid, userdetail.name, ttopuplog.topupdatetime, ttopuplog.topupvalue, 
            ttopuplog.oldcreditlimit, ttopuplog.newcreditlimit, 
            ttopuplog.topupby, ttopuplog.paylater, topuprequesttype.typedesc, topuprequest.notes 
            from userdetail, ttopuplog, topuprequest, topuprequesttype 
            where userdetail.accountid in (select accountid from userdetailext where wholesalerkey = {0}) 
            and userdetail.accountid = ttopuplog.accountid 
            and ttopuplog.topuprequestkey = topuprequest.indexkey 
            and topuprequest.topuptype = topuprequesttype.topuptype
            and ttopuplog.topupdatetime >= ? and ttopuplog.topupdatetime < ?
            order by ttopuplog.accountid, ttopuplog.topupdatetime
            """.format(wholesalerkey)
            
    return q

def topupreport_excel_list(datefrom, dateto, wholesalerkey=0, idxlist=None):
    b = None
    db = None
    
    try:
        wb = Workbook()
        ws = wb.active
        
        db = initdb()
        
        df = datetime.strptime(datefrom, '%Y-%m-%d')
        dt = datetime.strptime(dateto, '%Y-%m-%d')
        dk = dt + relativedelta(days=1)
        total = 0
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 38
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['G'].width = 13
        ws.column_dimensions['H'].width = 80
        
        if idxlist is None and wholesalerkey == 0:
            ws['A1'] = 'Customer Topup Report as at: {0}'.format(datetime.now().strftime('%d-%b-%Y'))
            ws.merge_cells('A1:C1')
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='right')
            
            ws['A2'] = 'Filter Date: {0} To {1}'.format(df.strftime('%d/%m/%Y'), dt.strftime('%d/%m/%Y'))
            ws['A2'].font = Font(size=12, bold=True)
            
            ws['A3'] = 'Filter Type: All customer'
            ws['A3'].font = Font(size=12, bold=True)
            
            lh = ['AccountID', 'Name', 'Topup DateTime', 'Topup Value', 'Login ID', 'PayLater', 'TopupType', 'Remark']
            j = 1
            for h in lh:
                ws.cell(row=4, column=j).value = h
                ws.cell(row=4, column=j).font = Font(bold=True)
                j += 1
                
            i = 5
            q = topupreport_query()
            rows = db.cur.execute(q, df, dk).fetchall()
            for r in rows:
                j = 1
                ws.cell(row=i, column=j).value = r.accountid
                
                j += 1
                ws.cell(row=i, column=j).value = r.name
                
                j += 1
                ws.cell(row=i, column=j).value = r.topupdatetime.strftime('%d/%m/%y %I:%M:%S %p')
                
                j += 1
                ws.cell(row=i, column=j).value = r.topupvalue
                
                j += 1
                ws.cell(row=i, column=j).value = r.topupby
                
                j += 1
                ws.cell(row=i, column=j).value = r.paylater
                
                j += 1
                ws.cell(row=i, column=j).value = r.typedesc
                
                j += 1
                ws.cell(row=i, column=j).value = r.notes
                
                i += 1
                
                total += r.topupvalue
                
        else:
            # Print report Header
            ws['A1'] = 'Customer Topup Report as at: {0}'.format(datetime.now().strftime('%d-%b-%Y'))
            ws.merge_cells('A1:C1')
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='right')
            
            ws['A2'] = 'Filter Date: {0} To {1}'.format(df.strftime('%d/%m/%Y'), dt.strftime('%d/%m/%Y'))
            ws['A2'].font = Font(size=12, bold=True)
            
            lh = ['AccountID', 'Name', 'Topup DateTime', 'Topup Value', 'Login ID', 'PayLater', 'TopupType', 'Remark']
            j = 1
            for h in lh:
                ws.cell(row=3, column=j).value = h
                ws.cell(row=3, column=j).font = Font(bold=True)
                j += 1
                
            i = 4
            q = topupreport_query_selected(idxlist, wholesalerkey)
            rows = db.cur.execute(q, df, dk).fetchall()
            for r in rows:
                j = 1
                ws.cell(row=i, column=j).value = r.accountid
                
                j += 1
                ws.cell(row=i, column=j).value = r.name
                
                j += 1
                ws.cell(row=i, column=j).value = r.topupdatetime.strftime('%d/%m/%y %I:%M:%S %p')
                
                j += 1
                ws.cell(row=i, column=j).value = r.topupvalue
                
                j += 1
                ws.cell(row=i, column=j).value = r.topupby
                
                j += 1
                ws.cell(row=i, column=j).value = r.paylater
                
                j += 1
                ws.cell(row=i, column=j).value = r.typedesc
                
                j += 1
                ws.cell(row=i, column=j).value = r.notes
                
                i += 1
                
                total += r.topupvalue
                
        ws.cell(row=i, column=3).value = 'Total Topup'
        ws.cell(row=i, column=4).value = total

        b = save_virtual_workbook(wb)
        
    finally:
        if db is not None:
            db.dispose()
            
    return b

def topupreport_text_list(datefrom, dateto, wholesalerkey=0, idxlist=None):
    b = None
    o = None
    db = None
    
    try:
        o = StringIO()
        
        db = initdb()
        
        df = datetime.strptime(datefrom, '%Y-%m-%d')
        dt = datetime.strptime(dateto, '%Y-%m-%d')
        dk = dt + relativedelta(days=1)
        total = 0
        grandtotal = 0
        
        if idxlist is None and wholesalerkey == 0:
            s = 'Customer Topup Report as at: {0}\r\n'.format(datetime.now().strftime('%d-%b-%Y'))
            o.write(s)
            
            s = 'Filter Date: {0} To {1}\r\n'.format(df.strftime('%d/%m/%Y'), dt.strftime('%d/%m/%Y'))
            o.write(s)
            o.write('Filter Type: All customer\r\n')
            o.write('==============================================================================\r\n')
            
            q = topupreport_query()
            rows = db.cur.execute(q, df, dk).fetchall()
            for r in rows:
                s = r.topupdatetime.strftime('%d-%m-%Y %H:%M:%S')
                o.write(s)
                o.write(''.ljust(22 - len(s)))
                
                s = '{0:6.2f}'.format(r.topupvalue)
                o.write(''.ljust(9 - len(s)))
                o.write(s)
                o.write('   ')
                
                s = '{0:8.2f}'.format(r.oldcreditlimit)
                o.write(''.ljust(11 - len(s)))
                o.write(s)
                o.write('   ')
                
                s = '{0:8.2f}'.format(r.newcreditlimit)
                o.write(''.ljust(11 - len(s)))
                o.write(s)
                o.write('   ')
                o.write(r.topupby[:15])
                
                if len(r.topupby[:15]) < 15:
                    o.write(''.ljust(15 - len(r.topupby[:15])))
                    o.write(' ')
                    
                o.write(r.typedesc[:20])
                o.write(''.ljust(20 - len(r.typedesc[:20])))
                
                o.write(r.notes)
                
                total += r.topupvalue
                grandtotal += r.topupvalue
                
                o.write('\r\n')
                
            o.write('           Total      ---------\r\n')
            s = '{0:6.2f}'.format(total)
            o.write('                      ' + (''.ljust(9 - len(s))) + s + '\r\n')
            
        else:
            j = 0
            dic = {}
            
            s = 'Customer Topup Report as at: {0}\r\n'.format(datetime.now().strftime('%d-%b-%Y %I:%M:%S %p'))
            o.write(s)
            
            s = 'Filter Date: {0} To {1}\r\n'.format(df.strftime('%d/%m/%Y'), dt.strftime('%d/%m/%Y'))
            o.write(s)
            o.write('==============================================================================\r\n')
            
            q = topupreport_query_selected(idxlist, wholesalerkey)
            rows = db.cur.execute(q, df, dk).fetchall()
            for r in rows:
                if r.accountid not in dic:
                    if len(dic) > 0:
                        o.write('           Total      ---------\r\n')
                        s = '{0:6.2f}'.format(total)
                        o.write('                      ' + (''.ljust(9 - len(s))) + s + '\r\n')
                        
                    o.write('\r\n')
                    dic[r.accountid] = 0
                    total = 0
                    s = '{0}.{1}-{2}\r\n'.format(j + 1, r.accountid, r.name)
                    o.write(s)
                    o.write('Topup Time               Amount   Old Balance   New Balance   PayLater Login ID       Topup Type           Remark\r\n')
                    o.write('-----------------------------------------------------------------------------------------------------------------\r\n')
                    j += 1
                    
                s = r.topupdatetime.strftime('%d-%m-%y %H:%M:%S')
                o.write(s)
                o.write(''.ljust(22 - len(s)))
                
                s = '{0:6.2f}'.format(r.topupvalue)
                o.write(''.ljust(9 - len(s)))
                o.write(s)
                o.write('   ')
                
                s = '{0:8.2f}'.format(r.oldcreditlimit)
                o.write(''.ljust(11 - len(s)))
                o.write(s)
                o.write('   ')
                
                s = '{0:8.2f}'.format(r.newcreditlimit)
                o.write(''.ljust(11 - len(s)))
                o.write(s)
                o.write('   ')
                
                # PayLater
                o.write('    ')
                o.write(str(r.paylater))
                o.write('    ')
                
                o.write(r.topupby[:15])
                
                if len(r.topupby[:15]) < 15:
                    o.write(''.ljust(15 - len(r.topupby[:15])))
                    o.write(' ')
                    
                o.write(r.typedesc[:20])
                o.write(''.ljust(20 - len(r.typedesc[:20])))
                o.write(' ')
                
                o.write(r.notes)
                
                total += r.topupvalue
                grandtotal += r.topupvalue
                
                o.write('\r\n')
                
            o.write('           Total      ---------\r\n')
            s = '{0:6.2f}'.format(total)
            o.write('                      ' + (''.ljust(9 - len(s))) + s + '\r\n')
            
        o.write('--------------------------------------------------\r\n')
        o.write('     Grand Total               {0:.2f}\r\n'.format(grandtotal))
        
        b = o.getvalue()
        
    finally:
        if db is not None:
            db.dispose()
            
        if o is not None:
            o.close()
            
    return b

def aninewreg_list_cust(keyword=''):
    l = []
    db = None
    
    try:
        db = initdb()
        
        if keyword == '':
            q = """
                select callerid, batchid, status, assigndate, assignee, blockdate, releasedate, code_area, state, wskey from GW_callerid
                order by assigndate 
                """
            db.cur.execute(q)
            # print(q)
            
        else:
            # q = """
            #     select accountid, name from userdetail 
            #     where accountid like ? or name like ? 
            #     order by name
            #     """
            q = """
                select callerid, batchid, status, assigndate, assignee, blockdate, releasedate, code_area, state,wskey from GW_callerid
                where callerid like ? or batchid like ? or state like ?
                order by assigndate 
                """
            print(keyword)
            db.cur.execute(q, '%{0}%'.format(keyword), '%{0}%'.format(keyword), '%{0}%'.format(keyword))
                
        rows = db.cur.fetchall()
        for r in rows:
            l.append({
                'callerid': r.callerid,
                'batchid': utils.decode(r.batchid),
                'status': r.status,
                'assigndate': r.assigndate,
                'assignee': r.assignee,
                'blockdate': r.blockdate,
                'releasedate': r.releasedate,
                'code_area': r.code_area,
                'state': r.state,
                'wskey': r.wskey,
            })
            
    finally:
        if db is not None:
            db.dispose()
            
    return l

def aninewreg_query(idxlist=None):
    if idxlist is None:
        q = """
            select a.name, b.callerid, b.accountid, b.creationdate, b.createdby 
            from userdetail a, authenticationex b 
            where a.accountid = b.accountid 
            and b.creationdate >= ? and b.creationdate < ? 
            order by b.accountid, b.creationdate, b.callerid
            """
            
    else:
        params = ','.join(["'{0}'".format(i) for i in idxlist])
        q = """
            select a.name, b.callerid, b.accountid, b.creationdate, b.createdby 
            from userdetail a, authenticationex b 
            where a.accountid = b.accountid 
            and b.creationdate >= ? and b.creationdate < ? 
            and b.accountid in ({0}) 
            order by b.accountid, b.creationdate, b.callerid
            """.format(params)
            
    return q

def aninewreg_excel_list(datefrom, dateto, idxlist=None):
    b = None
    db = None
    
    try:
        wb = Workbook()
        ws = wb.active
        
        db = initdb()
        
        df = datetime.strptime(datefrom, '%Y-%m-%d')
        dt = datetime.strptime(dateto, '%Y-%m-%d')
        dk = dt + relativedelta(days=1)
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 10
        
        ws['A1'] = 'ANI Status Change Report'
        ws.merge_cells('A1:C1')
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A2'] = 'Filter Date: {0} To {1}'.format(df.strftime('%d/%m/%Y'), dt.strftime('%d/%m/%Y'))
        ws['A2'].font = Font(size=12, bold=True)
        
        ws['A3'] = 'Filter Type: All customer'
        ws['A3'].font = Font(size=12, bold=True)
        
        lh = ['ANI', 'Last Update', 'Updated By']
        
        i = 4
        accountid = ''
        q = aninewreg_query(idxlist)
        rows = db.cur.execute(q, df, dk).fetchall()
        for r in rows:
            if accountid != r.accountid:
                i += 1
                # write customer accountID & name
                ws.cell(row=i, column=1).value = '{0}:{1}'.format(r.accountid, utils.decode(r.name))
                ws.cell(row=i, column=1).font = Font(size=14, bold=True)
                i += 1
                
                # write sub header
                j = 1
                for h in lh:
                    ws.cell(row=i, column=j).value = h
                    ws.cell(row=i, column=j).font = Font(bold=True, underline='single')
                    j += 1
                    
                i += 1
                accountid = r.accountid
                    
            ws.cell(row=i, column=1).value = utils.decode(r.callerid)
            ws.cell(row=i, column=2).value = r.creationdate.strftime('%d/%m/%Y %I:%M:%S %p')
            ws.cell(row=i, column=3).value = r.createdby
            i += 1

        b = save_virtual_workbook(wb)
        
    finally:
        if db is not None:
            db.dispose()
            
    return b

def aninewreg_text_list(datefrom, dateto, idxlist=None):
    l = []
    b = None
    o = None
    db = None
    
    try:
        o = StringIO()
        
        db = initdb()
        
        df = datetime.strptime(datefrom, '%Y-%m-%d')
        dt = datetime.strptime(dateto, '%Y-%m-%d')
        dk = dt + relativedelta(days=1)
        
        # Print report Header
        o.write('ANI Status Change Report\r\n')
        o.write('Filter Date: {0} To {1}\r\n'.format(df.strftime('%d/%m/%Y'), dt.strftime('%d/%m/%Y')))
        o.write('Filter Type: All customer\r\n')
        
        accountid = ''
        q = aninewreg_query(idxlist)
        rows = db.cur.execute(q, df, dk).fetchall()
        for r in rows:
            if accountid != r.accountid:
                o.write('\r\n')
                # write customer accountID & name
                o.write('{0}:{1}\r\n'.format(r.accountid, r.name))
                o.write('ANI{0}Last Update{1}Updated By\r\n'.format(''.ljust(15 - len('ANI')), ''.ljust(25 - len('Last Update'))))
                o.write('---------------------------------------------------------------\r\n')
                accountid = r.accountid
                
            creationdatestr = r.creationdate.strftime('%d/%m/%Y %I:%M:%S %p')
            o.write('{0}{1}{2}{3}{4}\r\n'.format(r.callerid, ''.ljust(15 - len(r.callerid)), 
                                                 creationdatestr, ''.ljust(25 - len(creationdatestr)),
                                                 r.createdby))
                
        b = o.getvalue()
        
    finally:
        if db is not None:
            db.dispose()
            
        if o is not None:
            o.close()
            
    return b

def ratetype_create(o):
    db = None
    
    try:
        assert isinstance(o, models.RateType)
        
        if o.iratetype == 0 or o.iratetype > 999:
            raise UIException('Invalid rate type')
        
        db = initdb(dbx=None, autocommit=False)
        
        q = """
            insert into tratetype (iratetype, sname, imaintype) 
            values (?, ?, ?)
            """
        db.cur.execute(q, o.iratetype, o.sname, 1)
        db.commit()
        
    except pyodbc.IntegrityError as e:
        if utils.isduplicatekey(e):
            raise UIException('Rate Type {0} already exist'.format(o.iratetype))
        
        else:
            raise UIException(traceback.format_exc())
        
    finally:
        if db is not None:
            db.dispose()
        
def ratetype_update(o):
    db = None
    
    try:
        assert isinstance(o, models.RateType)
        
        db = initdb(dbx=None, autocommit=False)
        
        q = 'update tratetype set sname = ? where iratetype = ?'
        db.cur.execute(q, o.sname, o.iratetype)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def userdetail_copy(srcaccountid, o, x, user):
    db = None
    
    try:
        assert isinstance(o, models.UserDetail)
        assert isinstance(x, models.UserDetailExt)
        
        srcaccountid = srcaccountid.zfill(9)
        if len(srcaccountid) != 9:
            raise UIException('Invalid accountid!! Account ID must be a 9 digits number')
        
        db = initdb(dbx=None, autocommit=False)
        
        k, name = isaccountid_exist(x.accountid, db)
        if k == False:
            raise UIException('Cannot find AccountID: {0} in the database'.format(x.accountid))
        
        k, name = isaccountid_exist(srcaccountid, db)
        if k == True:
            raise UIException('The new AccountID has been used by Customer: {0}'.format(name))
        
        q = """
            insert into userdetail
                       (accountid
                       ,name
                       ,address
                       ,creditlimit
                       ,creditusage
                       ,imaintype
                       ,isubtype
                       ,sfeaturescode
                       ,igatetype
                       ,iratetype
                       ,ilang)
            select ?
                  ,name
                  ,address
                  ,creditlimit
                  ,creditusage
                  ,imaintype
                  ,isubtype
                  ,sfeaturescode
                  ,igatetype
                  ,iratetype
                  ,ilang from userdetail 
                  where accountid = ?
            """
        db.cur.execute(q, srcaccountid, x.accountid)
        
        q = """
            insert into userdetailext (accountid, createdby, note, wholesalerkey, emailaddress, 
            addr1, addr2, addr3, city, state, 
            postcode, acstatus, acmanager, usagecategory, pbxmodel, 
            supportteam, technicalnotes, sipsubtype, siplcrtype, iddusagealert, 
            iddusagebar, mobusagealert, mobusagebar, stdusagealert, stdusagebar)
            values (?, ?, ?, ?, ?, 
            ?, ?, ?, ?, ?, 
            ?, ?, ?, ?, ?, 
            ?, ?, ?, ?, ?, 
            ?, ?, ?, ?, ?)
            """
        db.cur.execute(q, srcaccountid, user.username, utils.removequotation(x.note), x.wholesalerkey, utils.removequotation(x.emailaddress), 
                       utils.removequotation(x.addr1), utils.removequotation(x.addr2), utils.removequotation(x.addr3), utils.removequotation(x.city), utils.removequotation(x.state),
                       utils.removequotation(x.postcode), x.acstatus, utils.removequotation(x.acmanager), x.usagecategory, utils.removequotation(x.pbxmodel),
                       x.supportteam, utils.removequotation(x.technicalnotes), x.sipsubtype, x.siplcrtype, x.iddusagealert,
                       x.iddusagebar, x.mobusagealert, x.mobusagebar, x.stdusagealert, x.stdusagebar)
        
        q = 'select callerid, accountid, status from authentication where accountid = ?'
        rows = db.cur.execute(q, x.accountid).fetchall()
        la = []
        for r in rows:
            a = models.Authentication()
            a.callerid = r.callerid
            a.accountid = r.accountid
            a.status = r.status
            la.append(a)
            
        for a in la:
            q = """
                delete from authentication where callerid = ? and accountid = ?
                """
            db.cur.execute(q, a.callerid, a.accountid)
            
            q = """
                insert into authentication (callerid, accountid, status) 
                values (?, ?, ?)
                """
            db.cur.execute(q, a.callerid, srcaccountid, a.status)
            
        if utils.isvkiosktype(o.isubtype):
            q = """
                select accountid, callerid, surchargetype, smsno, faxno, phoneno 
                from vkNotification 
                where accountid = ?
                """
            r = db.cur.execute(q, x.accountid).fetchone()
            v = models.VkNotification()
            v.accountid = r.accountid
            v.callerid = r.callerid
            v.surchargetype = r.surchargetype
            v.smsno = r.smsno
            v.faxno = r.faxno
            v.phoneno = r.phoneno

            q = """
                delete from vkNotification where accountid = ?
                """
            db.cur.execute(q, x.accountid)
            
            q = """
                insert into vkNotification (accountid, callerid, surchargetype, smsno, faxno, phoneno) 
                values (?, ?, ?, ?, ?, ?)
                """
            db.cur.execute(q, srcaccountid, v.callerid, v.surchargetype, v.smsno, v.faxno, v.phoneno)
            
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def igatetype_update(o, u, db):
    db1 = None
    db2 = None
    
    try:
        assert isinstance(o, models.UserDetail)
        assert isinstance(u, models.UserDetail)
        
        if o.igatetype in [constants.GWID.GWALL, constants.GWID.GWCP]:
            raise UIException('Unable to change from CP gateway')
        
        elif o.igatetype in [constants.GWID.GWAM, constants.GWID.GWKLP, constants.GWID.GWPD, constants.GWID.GWSB]:
            db1, gw1 = utils.connectGWDB(o.igatetype, autocommit=False)
            
        if u.igatetype in [constants.GWID.GWALL, constants.GWID.GWCP]:
            raise UIException('Unable to change to CP gateway')
        
        elif u.igatetype in [constants.GWID.GWAM, constants.GWID.GWKLP, constants.GWID.GWPD, constants.GWID.GWSB]:
            db2, gw2 = utils.connectGWDB(u.igatetype, autocommit=False)
            
        db = initdb(dbx=None, autocommit=False)
            
        if db1 is not None:
            q = 'select igatetype, creditlimit, creditusage from userdetail where accountid = ?'
            r = db1.cur.execute(q, o.accountid).fetchone()
            if r is None:
                raise UIException('Cannot find account: {0} at {1} gateway'.format(o.accountid, gw1))
            
            if db2 is not None:
                a = models.UserDetail()
                a.igatetype = r.igatetype
                a.creditlimit = r.creditlimit
                a.creditusage = r.creditusage
                
                q = """
                    update userdetail set igatetype = ?, creditlimit = ?, creditusage = ? 
                    where accountid = ?
                    """
                db2.cur.execute(q, o.igatetype, a.creditlimit, a.creditusage, o.accountid)
            
            q = 'update userdetail set igatetype = ? where accountid = ?'
            db1.cur.execute(q, o.igatetype, o.accountid)
            
            q = """
                update userdetail set creditusage = ? 
                where accountid = ?
                """
            db.cur.execute(q, a.creditusage, o.accountid)

            if db1 is not None:
                db1.commit()

            if db2 is not None:
                db2.commit()

        db.commit()
        
    finally:
        if db1 is not None:
            db1.dispose()
            
        if db2 is not None:
            db2.dispose()

def userdetail_create(o, x, user):
    db = None
    
    try:
        assert isinstance(o, models.UserDetail)
        assert isinstance(x, models.UserDetailExt)
        
        db = initdb(dbx=None, autocommit=False)
        
        # Check accountid ID Exists
        k, name = isaccountid_exist(o.accountid, db)
        if k:
            raise UIException('Account ID Already Exists In The Database...')
        
        k, accountid = isname_exist(o.name, db)
        if k:
            raise UIException('Customer Name Already Exists In The Database...')
        
        checktype(o)
        
        q = """
            insert into userdetail (accountid, name, address, creditlimit, creditusage, imaintype, isubtype, sfeaturescode, igatetype, iratetype, ilang) 
            values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
        db.cur.execute(q, o.accountid, utils.removequotation(o.name), utils.removequotation(o.address), 0, 0, 
                       o.imaintype, o.isubtype, '', o.igatetype, o.iratetype, 
                       '{0}{1}'.format(o.ilcrtype, o.ilang))
        
        q = """
            insert into userdetailext (accountid, createdby, note, wholesalerkey, emailaddress, 
            addr1, addr2, addr3, city, state, 
            postcode, acstatus, acmanager, usagecategory, pbxmodel, 
            supportteam, technicalnotes, sipsubtype, siplcrtype, iddusagealert, 
            iddusagebar, mobusagealert, mobusagebar, stdusagealert, stdusagebar)
            values (?, ?, ?, ?, ?, 
            ?, ?, ?, ?, ?, 
            ?, ?, ?, ?, ?, 
            ?, ?, ?, ?, ?, 
            ?, ?, ?, ?, ?)
            """
        db.cur.execute(q, x.accountid, user.username, utils.removequotation(x.note), x.wholesalerkey, utils.removequotation(x.emailaddress), 
                       utils.removequotation(x.addr1), utils.removequotation(x.addr2), utils.removequotation(x.addr3), utils.removequotation(x.city), utils.removequotation(x.state),
                       utils.removequotation(x.postcode), x.acstatus, utils.removequotation(x.acmanager), x.usagecategory, utils.removequotation(x.pbxmodel),
                       x.supportteam, utils.removequotation(x.technicalnotes), x.sipsubtype, x.siplcrtype, x.iddusagealert,
                       x.iddusagebar, x.mobusagealert, x.mobusagebar, x.stdusagealert, x.stdusagebar)
        db.commit()
        
#         qs = """
#              if not exists (select AccountID from UserDetailExt where AccountID = ?) 
#              insert into UserDetailExt (AccountID, SipSubType, SipLcrType, IDDUsageBar, IDDUsageAlert, MOBUsageBar, MOBUsageAlert, STDUsageBar, STDUsageAlert) 
#              values ('{0}', {1}, {2}, {3}, {4}, {5}, {6}, {7}, {8})
#              """.format(x.accountid, x.sipsubtype, x.siplcrtype, x.iddusagebar, x.iddusagealert, x.mobusagebar, x.mobusagealert, x.stdusagebar, x.stdusagealert)
#         insert_into_sql_log(qs, x.accountid, username, db, 0)
        
    finally:
        if db is not None:
            db.dispose()

def userdetail_update(o, x, y, dic, user, pwd, confirm=False):
    db = None
    prompt = None
    
    try:
        assert isinstance(o, models.UserDetail)
        assert isinstance(x, models.UserDetailExt)
        assert isinstance(y, models.MaxCallAppearanceDN)
        #y = models.MaxCallAppearanceDN()
        
        # Check Record Error
        utils.checkemptyuserdetail(dic)
        
        db = initdb(dbx=None, autocommit=False, user=user.username, pwd=pwd)
        
        u = userdetail_get(o.accountid, "", db)
        assert isinstance(u, models.UserDetail)

        # check is acstatus = suspend & terminated
        if utils.isadmin(user.functionflag) == False:
            if u.userdetailext.acstatus in [2, 3]:
                raise UIException('You cannot change the account because it is suspended/terminated, please update the account status first')
            
        # Check Name Exists
        k, accountid = isname_exist(o.name, db)
        if k and o.name.strip() != u.name.strip():
            raise UIException('Customer Name Already Exists In The Database...')
        
        da = utils.checkaccesslevel(user.iaccesslevel)
        if o.igatetype != u.igatetype and constants.ACCESS.CHANGEGATETYPE not in da:
            if o.isubtype >= 600 and o.isubtype < 700:
                if confirm == False:
                    prompt = 'You are changing gateway type. Are you sure want to do that ?'
                
                else:
                    igatetype_update(o, u, db)
                    
        if prompt is None:
            # Update User Detail
            if o.igatetype == u.igatetype and o.name == u.name and u.igatetype in [6, 7]:
                q = """
                    update userdetail set name = ?, address = ?, imaintype = ?, isubtype = ?, iratetype = ?, 
                    igatetype = ?, ilang = ?, pbxno = ?  
                    where accountid = ?
                    """
                db.cur.execute(q, utils.removequotation(o.name), utils.removequotation(o.address), o.imaintype, o.isubtype, o.iratetype,
                               o.igatetype, '{0}{1}'.format(o.ilcrtype, o.ilang), o.pbxno, o.accountid)                  
            else:
                q = """
                    update userdetail set name = ?, address = ?, imaintype = ?, isubtype = ?, iratetype = ?, 
                    igatetype = ?, ilang = ?, pbxno = ?  
                    where accountid = ?
                    """
                db.cur.execute(q, utils.removequotation(o.name), utils.removequotation(o.address), o.imaintype, o.isubtype, o.iratetype,
                               o.igatetype, '{0}{1}'.format(o.ilcrtype, o.ilang), o.pbxno, o.accountid)
               
            mca = u.maxcallappearancedn
            if o.igatetype == u.igatetype and o.name == u.name and u.igatetype == 7:
                #AFF
                print(u.igatetype)
                print(mca.MaxCallAppearance)
                if mca.MaxCallAppearance != y.MaxCallAppearance:
                    qi = "update MaxCallAppearanceDN set MaxCallAppearance = ? where DirectoryNumber = ?"
                    print(y.MaxCallAppearance)
                    db.cur.execute(qi, y.MaxCallAppearance, o.pbxno)

                #else:    
                if y.MaxCallAppearance is None:
                    qi = """
                        insert into MaxCallAppearanceDN (DirectoryNumber, MaxCallAppearance )
                        values(?,1)
                        """   
                    db.cur.execute(qi,o.pbxno)
                else: 
                    qi = 'delete from MaxCallAppearanceDN where DirectoryNumber = ?'
                    db.cur.execute(qi,o.pbxno)

                    qi = """
                        insert into MaxCallAppearanceDN (DirectoryNumber, MaxCallAppearance )
                        values(?,?)
                        """   
                    db.cur.execute(qi,o.pbxno,y.MaxCallAppearance)
                    #print(qi)

            uext = u.userdetailext
            
            if uext.note != utils.removequotation(x.note):
                qi = "update userdetailext set note = ? where accountid = ?"
                db.cur.execute(qi, utils.removequotation(x.note), o.accountid)

            if uext.wholesalerkey != x.wholesalerkey:
                qi = "update userdetailext set wholesalerkey = ? where accountid = ?"
                db.cur.execute(qi, x.wholesalerkey, o.accountid)

            if uext.status != x.status:
                qi = "update userdetailext set status = ? where accountid = ?"
                print(qi)
                db.cur.execute(qi, x.status, o.accountid)
                qs = "UPDATE UserDetailExt SET Status={0} WHERE AccountID='{1}'".format(x.status, o.accountid)
                insert_into_sql_log(qs, o.accountid, user.username, db, -1)
                #update_status_wscallbilling(x.status, o.accountid)
                qii = "update UserDetailExt_dup set status = ? where accountid = ?"
                # print(qii)
                db.cur.execute(qii, x.status, o.accountid)
                    
               

            if uext.emailaddress != utils.removequotation(x.emailaddress):
                qi = "update userdetailext set emailaddress = ? where accountid = ?"
                db.cur.execute(qi, utils.removequotation(x.emailaddress), o.accountid)
            
            if uext.addr1 != utils.removequotation(x.addr1):
                qi = "update userdetailext set addr1 = ? where accountid = ?"
                db.cur.execute(qi, utils.removequotation(x.addr1), o.accountid)

            if uext.addr2 != utils.removequotation(x.addr2):
                qi = "update userdetailext set addr2 = ? where accountid = ?"
                db.cur.execute(qi, utils.removequotation(x.addr2), o.accountid)

            if uext.addr3 != utils.removequotation(x.addr3):
                qi = "update userdetailext set addr3 = ? where accountid = ?"
                db.cur.execute(qi, utils.removequotation(x.addr3), o.accountid)

            if uext.city != utils.removequotation(x.city):
                qi = "update userdetailext set city = ? where accountid = ?"
                db.cur.execute(qi, utils.removequotation(x.city), o.accountid)

            if uext.state != x.state:
                qi = "update userdetailext set state = ? where accountid = ?"
                db.cur.execute(qi, x.state, o.accountid)

            if uext.postcode != utils.removequotation(x.postcode):
                qi = "update userdetailext set postcode = ? where accountid = ?"
                db.cur.execute(qi, utils.removequotation(x.postcode), o.accountid)

            if uext.acstatus != x.acstatus:
                qi = "update userdetailext set acstatus = ? where accountid = ?"
                db.cur.execute(qi, x.acstatus, o.accountid)

            if uext.acmanager != utils.removequotation(x.acmanager):
                qi = "update userdetailext set acmanager = ? where accountid = ?"
                db.cur.execute(qi, utils.removequotation(x.acmanager), o.accountid)

            if uext.usagecategory != x.usagecategory:
                qi = "update userdetailext set usagecategory = ? where accountid = ?"
                db.cur.execute(qi, x.usagecategory, o.accountid)

            if uext.pbxmodel != utils.removequotation(x.pbxmodel):
                qi = "update userdetailext set pbxmodel = ? where accountid = ?"
                db.cur.execute(qi, utils.removequotation(x.pbxmodel), o.accountid)

            if uext.supportteam != x.supportteam:
                qi = "update userdetailext set supportteam = ? where accountid = ?"
                db.cur.execute(qi, x.supportteam, o.accountid)

            if uext.technicalnotes != utils.removequotation(x.technicalnotes):
                qi = "update userdetailext set technicalnotes = ? where accountid = ?"
                db.cur.execute(qi, utils.removequotation(x.technicalnotes), o.accountid)
            
            if uext.sipsubtype != x.sipsubtype:
                qi = "update userdetailext set sipsubtype = ? where accountid = ?"
                db.cur.execute(qi, x.sipsubtype, o.accountid)

            if uext.siplcrtype != x.siplcrtype:
                qi = "update userdetailext set siplcrtype = ? where accountid = ?"
                db.cur.execute(qi, x.siplcrtype, o.accountid)

            if uext.iddusagealert != x.iddusagealert:
                qi = "update userdetailext set iddusagealert = ? where accountid = ?"
                db.cur.execute(qi, x.iddusagealert, o.accountid)

            if uext.iddusagebar != x.iddusagebar:
                qi = "update userdetailext set iddusagebar = ? where accountid = ?"
                db.cur.execute(qi, x.iddusagebar, o.accountid)

            if uext.mobusagealert != x.mobusagealert:
                qi = "update userdetailext set mobusagealert = ? where accountid = ?"
                db.cur.execute(qi, x.mobusagealert, o.accountid)

            if uext.mobusagebar != x.mobusagebar:
                qi = "update userdetailext set mobusagebar = ? where accountid = ?"
                db.cur.execute(qi, x.mobusagebar, o.accountid)

            if uext.stdusagealert != x.stdusagealert:
                qi = "update userdetailext set stdusagealert = ? where accountid = ?"
                db.cur.execute(qi, x.stdusagealert, o.accountid)
            
            if uext.stdusagebar != x.stdusagebar:
                qi = "update userdetailext set stdusagebar = ? where accountid = ?"
                db.cur.execute(qi, x.stdusagebar, o.accountid)

            
            # q = """
            #     update userdetailext set note = ?, wholesalerkey = ?, status = ?, emailaddress = ?, addr1 = ?, 
            #     addr2 = ?, addr3 = ?, city = ?, state = ?, postcode = ?, 
            #     acstatus = ?, acmanager = ?, usagecategory = ?, pbxmodel = ?, supportteam = ?, 
            #     technicalnotes = ?, sipsubtype = ?, siplcrtype = ?, iddusagealert = ?, iddusagebar = ?, 
            #     mobusagealert = ?, mobusagebar = ?, stdusagealert = ?, stdusagebar = ? 
            #     where accountid = ?
            #     """
            # db.cur.execute(q, x.note, x.wholesalerkey, x.status, x.emailaddress, x.addr1,
            #                x.addr2, x.addr3, x.city, x.state, x.postcode,
            #                x.acstatus, x.acmanager, x.usagecategory, x.pbxmodel, x.supportteam,
            #                x.technicalnotes, x.sipsubtype, x.siplcrtype, x.iddusagealert, x.iddusagebar,
            #                x.mobusagealert, x.mobusagebar, x.stdusagealert, x.stdusagebar,
            #                o.accountid)
            db.commit()
        
    finally:
        if db is not None:
            db.dispose()
            
    return prompt

def userdetail_delete(accountid, user, pwd):
    le = []
    db = None
    
    try:
        ex = UIException('Data not Complete (Customer AccountID)...')
        
        if accountid is None:
            raise ex
        
        else:
            accountid = accountid.strip()
            if accountid == '':
                raise ex
        
        db = initdb(user=user, pwd=pwd)
        
        # Check accountid ID Exists
        k, name = isaccountid_exist(accountid, db)
        if k == False:
            raise UIException('Cannot find record in the database')
        
        q = 'delete from userdetail where accountid = ?'
        db.cur.execute(q, accountid)
        
        q = 'delete from userdetailext where accountid = ?'
        db.cur.execute(q, accountid)
        
        q = 'delete from authentication where accountid = ?'
        db.cur.execute(q, accountid)
        
        q = 'delete from userdetailremark where rmk_accountid = ?'
        db.cur.execute(q, accountid)
        
        lbb = bbauthentication_list(accountid, db)
        
        for o in lbb:
            ex = bbauthentication_delete_callerid(db, o.bb_rt015, 0)
            if ex is not None:
                le.append(ex)
                
    finally:
        if db is not None:
            db.dispose()
                
    return le

def isaccountid_exist(accountid, dbx=None):
    b = False
    name = None
    db = dbx
    
    try:
        db = initdb(dbx)
    
        q = 'select accountid, name from userdetail where accountid = ?'
        r = db.cur.execute(q, accountid).fetchone()
        if r is not None:
            b = True
            name = r.name
            
    finally:
        if db is not None and dbx is None:
            db.dispose()
        
    return b, name

def isname_exist(name, dbx=None):
    b = False
    accountid = None
    db = dbx
    
    try:
        db = initdb(dbx)
    
        q = 'select name, accountid from userdetail where name = ?'
        r = db.cur.execute(q, name).fetchone()
        if r is not None:
            b = True
            accountid = r.accountid
            
    finally:
        if db is not None and dbx is None:
            db.dispose()
        
    return b, accountid

def checktype(o):
    assert isinstance(o, models.UserDetail)
    
    if o.isubtype >= 600 and o.isubtype < 700:
        if o.igatetype != constants.GWID.GWAM and o.igatetype != constants.GWID.GWSB:
            raise UIException('Prepaid type customer can only upload to AIM or SG gateway')
        
        if o.iratetype == 0:
            raise UIException('Prepaid Subtype cannot assign to postpaid ratetype')
        
    if utils.iscallshoptype(o.isubtype) and \
        o.igatetype != constants.GWID.GWAM and o.igatetype != constants.GWID.GWKLP:
        raise UIException('Call Shop customer can only assign to PD (02) or KLP(04) gateway')
    
    if utils.iscallshoptype(o.isubtype) and o.iratetype == 0:
        raise UIException('Call Shop prepaid customer cannot assign to postpaid ratetype')
    
    if o.isubtype == constants.VIRTUALKIOSKTYPE:
        if o.iratetype < 500 or o.iratetype > 599:
            raise UIException("Wrong virtual kiosk customer's ratetype")
        
    if checkbudgetcall(o) == False:
        return
    
    if checklcrtype(o) == False:
        return
        
def checkbudgetcall(o):
    assert isinstance(o, models.UserDetail)
    
    if isbudgetcallcustomer(o.accountid):
        if o.ilcrtype != constants.BUDGETLCR:
            raise UIException('Budget call customer must select budget LCR type')
        
        if o.iratetype != constants.BUDGETRATE:
            raise UIException('Budget call customer must select budget call rate type')
        
    else:
        if o.ilcrtype == constants.BUDGETLCR:
            raise UIException('Only Budget call customer can select budget LCR type')
        
        if o.iratetype == constants.BUDGETRATE:
            raise UIException('Only Budget call customer can select budget call rate type')

def checklcrtype(o):
    assert isinstance(o, models.UserDetail)
    
    firstdigit = o.accountid[0]
    first3digit = o.accountid[:3]
    
    if firstdigit in ['7', '4'] and o.ilcrtype != 5:
        raise UIException('Kiosk customer(accountid start with 7 or 5) can only using LCR Type-5')
    
    if first3digit == '881' and o.ilcrtype != 5:
        raise UIException('Kiosk customer(accountid start with 881) can only using LCR Type-5')
    
def isbudgetcallcustomer(accountid):
    b = False
    
    if len(accountid) < 5:
        return b
    
    code = accountid[3:5]
    return b

# not used
def iskioskcrosspayphone(callerid, subtype):
    b = False
    db = None
    dbx = None
    
    return b
    
    try:
        if utils.isminikiosk(subtype) == False:
            return b
        
        dbx, gw = utils.connectGWDB(constants.GWID.GWPDDB2)
        
        q = 'select cardno, callerid from MCwCLID where callerid = ?'
        rx = dbx.cur.execute(q, callerid).fetchone()
        if rx is not None:
            cardno = rx.cardno[:7]
            dbx.dispose()
            dbx = None
            
            db = initdb()
            
            q = 'select * from motherCcard where pin = ? and wsid = ?'
            r = db.cur.execute(q, cardno, '001170000').fetchone()
            if r is not None:
                b = True
                
        if b:
            raise UIException('You cannot register CallerID {0} for this Kiosk/MiniKiosk customer because it has been registered by PayPhone Customer'.format(callerid))
                
    finally:
        if dbx is not None:
            dbx.dispose()
            
        if db is not None:
            db.dispose()
            
    return b

def tlogin_password_get(username):
    s = None
    db = None

    try:
        db = initdb()

        q = """
            select sloginname, sloginpassword from tlogin 
            where sloginname = ?
            """
        r = db.cur.execute(q, username).fetchone()
        if r is not None:
            s = utils.simpledecrypt(r.sloginpassword, 8)

    finally:
        if db is not None:
            db.dispose()

    return s

def tlogin_get(username, password):
    o = None
    db = None
    
    try:
        db = initdb(user=username, pwd=password)
        
        q = """
            select sloginname, sloginpassword, iaccesslevel, ip, itopupaccesslevel, twholesaler, featureflag, 
            functionflag, department, admin, expiry from tlogin 
            where sloginname = ?
            """
        r = db.cur.execute(q, username).fetchone()
        if r is None:
            raise Exception('User {0} not found'.format(username))
        
        o = models.TLogin()
        o.sloginname = r.sloginname
        o.sloginpassword = r.sloginpassword
        o.iaccesslevel = r.iaccesslevel
        o.ip = r.ip
        o.itopupaccesslevel = r.itopupaccesslevel
        o.twholesaler = r.twholesaler
        o.featureflag = r.featureflag
        o.functionflag = r.functionflag
        o.department = r.department
        o.admin = r.admin
        o.expiry = r.expiry
        
    finally:
        if db is not None:
            db.dispose()
            
    return o

def changepassword(o):
    db = None
    
    try:
        assert isinstance(o, AccountRenew)
        o.password = o.password.strip()
        o.newpassword = o.newpassword.strip()
        
        db = initdb(user=o.username, pwd=o.password)
        
        q = """
            SET NOCOUNT ON
            declare
              @return_value int
            exec @return_value = sp_ChangePassword_GoldenKey
              @oldPsw = ?,
              @newPsw = ?,
              @LoginID = ?
            select
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, o.password, o.newpassword, o.username).fetchone()
        if r.Ret != 0:
            raise Exception('Failed to change password')
        
        q = 'update tlogin set sloginpassword = ?, expiry = ? where sloginname = ?'
        encpwd = utils.simpleencrypt(o.newpassword, 8)
        dt = datetime.now()
        dx = dt + relativedelta(months=3)
        db.cur.execute(q, encpwd, dx, o.username)
        
        q = 'update mrplogin SET sloginpassword = ?, expiry = ? where sloginname = ?'
        db.cur.execute(q, encpwd, dx, o.username)
        
    finally:
        if db is not None:
            db.dispose()
            
def topupstatus_list(wholesalerkey, datefrom):
    l = []
    db = None
    
    try:
        df = datetime.strptime(datefrom, '%Y-%m-%d')
        dk = df + relativedelta(days=1)
        
        db = initdb()
        
        q = """
            select a.accountid, a.topupdatetime, a.topupvalue, a.topupby, b.igatetype 
            from ttopuplog a, userdetail b 
            where topupdatetime >= ? and topupdatetime < ? 
            and a.accountid = b.accountid 
            and a.accountid in (select accountid from userdetailext where wholesalerkey = ?)
            """
        rows = db.cur.execute(q, df, dk, wholesalerkey).fetchall()
        for r in rows:
            o = models.TopupStatus()
            o.accountid = r.accountid
            o.topupdatetime = r.topupdatetime
            o.topupvalue = r.topupvalue
            o.topupby = r.topupby
            o.igatetype = r.igatetype
            
            qx = """
                 select status, uploaddate, accountid, uploadby 
                 from gwlog where accountid = ? and transactionsql like ? 
                 and (status & ?) <> 0 
                 and creationdate <= ? and creationdate >= ? 
                 """
            dc = r.topupdatetime + relativedelta(minutes=-2)
            db.cur.execute(qx, r.accountid, 'UPDATE UserDetail SET CreditLimit=%', o.igatetype, o.topupdatetime, dc)
            rx = db.cur.fetchone()
            if rx is not None:
                o.uploaddate = rx.uploaddate
                o.uploadby = rx.uploadby
                
            l.append(o)
        
    finally:
        if db is not None:
            db.dispose()
    
    return l

def wholesaler_acstatus_suspend(wholesalerkey):
    db = None

    try:
        db = initdb(dbx=None, autocommit=False)

        q = """
            update userdetailext set acstatus = 2 
            where wholesalerkey = ? and acstatus = 1
            """
        db.cur.execute(q, wholesalerkey)
        db.commit()

    finally:
        if db is not None:
            db.dispose()

def wholesaler_acstatus_reactivate(wholesalerkey):
    db = None

    try:
        db = initdb(dbx=None, autocommit=False)

        q = """
            update userdetailext set acstatus = 1 
            where wholesalerkey = ? and acstatus = 2
            """
        db.cur.execute(q, wholesalerkey)
        db.commit()

    finally:
        if db is not None:
            db.dispose()

def cust_sync_balance(accountid, igatetype, creditlimit):
    dbx = None
    db = None
    m = {}

    try:
        if igatetype in [constants.GWID.GWALL, constants.GWID.GWCP]:
            raise UIException('Unable to refresh the credit balance of customer who resides at CP gateway')

        dbx, gw = utils.connectGWDB(igatetype)

        q = 'select isubtype, creditlimit, creditusage from userdetail where accountid = ?'
        rx = dbx.cur.execute(q, accountid).fetchone()
        if rx is None:
            raise UIException('Cannot find account: {0} at {1} gateway'.format(accountid, gw))

        newlimit = rx.creditlimit
        newusage = float(rx.creditusage)

        db = initdb(autocommit=False)

        q = 'update userdetail set creditusage = ? where accountid = ?'
        db.cur.execute(q, newusage, accountid)
        db.commit()

        m['creditusage'] = newusage
        m['creditbalance'] = float(creditlimit) - newusage

    finally:
        if dbx is not None:
            dbx.dispose()

        if db is not None:
            db.dispose()

    return m

# def view015numbers_excel_list(wskey):
#     b = None
#     db = None

#     try:
#         db = initdb()

#         wb = Workbook()
#         ws = wb.create_sheet('Sheet1', 0)

#         ws.column_dimensions['A'].width = 18
#         ws.column_dimensions['B'].width = 22
#         ws.column_dimensions['D'].width = 13
#         ws.column_dimensions['E'].width = 13
#         lh = ['RT 015/03 Number', 'Creation Date', 'Status', 'Nice Number', 'Wholesaler']
#         j = 1
#         for h in lh:
#             ws.cell(row=1, column=j).value = h
#             j += 1

#         i = 2
#         q = """
#             select bb_rt015, bb_creationdate, bb_status, bb_nicenum, bb_wskey from bb_authentication 
#             where bb_wskey = ? 
#             order by bb_rt015
#             """
#         rows = db.cur.execute(q, wskey).fetchall()
#         for r in rows:
#             ws.cell(row=i, column=1).value = r.bb_rt015
#             ws.cell(row=i, column=2).value = r.bb_creationdate
#             ws.cell(row=i, column=3).value = r.bb_status
#             ws.cell(row=i, column=4).value = r.bb_nicenum
#             ws.cell(row=i, column=5).value = r.bb_wskey
#             i += 1

#         ws = wb.create_sheet('Sheet2', 1)

#         ws.column_dimensions['A'].width = 18
#         ws.column_dimensions['B'].width = 13
#         ws.column_dimensions['C'].width = 22
#         ws.column_dimensions['D'].width = 35
#         lh = ['RT 015/03 Number', 'Account ID', 'Creation Date', 'Wholesaler']
#         j = 1
#         for h in lh:
#             ws.cell(row=1, column=j).value = h
#             j += 1

#         i = 2
#         q = """
#             select a.callerid, a.accountid, aex.creationdate, c.wholesalername 
#             from authentication a left join userdetailext b on a.accountid = b.accountid
#             left join authenticationex aex on a.callerid = aex.callerid 
#             left join twholesaler c on b.wholesalerkey = c.wholesalerkey
#             where b.wholesalerkey = ? and a.callerid in (
#                 select bb_rt015 from bb_authentication where bb_wskey = ?
#             ) 
#             order by a.callerid
#             """
#         rows = db.cur.execute(q, wskey, wskey).fetchall()
#         for r in rows:
#             ws.cell(row=i, column=1).value = r.callerid
#             ws.cell(row=i, column=2).value = r.accountid
#             ws.cell(row=i, column=3).value = r.creationdate
#             ws.cell(row=i, column=4).value = r.wholesalername
#             i += 1

#         b = save_virtual_workbook(wb)

#     finally:
#         if db is not None:
#             db.dispose()

#     return b


def view015numbers_excel_list(wskey):
    
    print(wskey)
    b = None
    db = None

    try:
        db = initdb()

        wb = Workbook()
        ws = wb.create_sheet('Available Number', 0)

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 22
        ws.column_dimensions['H'].width = 13
        ws.column_dimensions['I'].width = 13
        ws.column_dimensions['J'].width = 13


        lh = ['Batch ID','RT 015/03 Number', 'Status', 'Wholesaler','Creation Date','Block Date', 'Release Date', 'State', 'Charge Area'   ]
        j = 1
        for h in lh:
            ws.cell(row=1, column=j).value = h
            j += 1

        i = 2
        q = """
            select batchid, callerid, status, wskey, assigndate, blockdate, releasedate, code_area, state from GW_callerid
            where status=0 and wskey = ? 
            order by assigndate 
        
            """
        # db.cur.execute(q, '%{0}%'.format(batchid))
        # rows = db.cur.fetchall()

        rows = db.cur.execute(q, wskey).fetchall()
        for r in rows:
            ws.cell(row=i, column=1).value = r.batchid
            ws.cell(row=i, column=2).value = r.callerid
            ws.cell(row=i, column=3).value = r.status
            ws.cell(row=i, column=4).value = r.wskey
            ws.cell(row=i, column=5).value = r.assigndate
            ws.cell(row=i, column=6).value = r.blockdate
            ws.cell(row=i, column=7).value = r.releasedate
            ws.cell(row=i, column=8).value = r.code_area
            ws.cell(row=i, column=9).value = r.state
   
            i += 1

        ws = wb.create_sheet('Used Number', 1)

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 22
        ws.column_dimensions['H'].width = 13
        ws.column_dimensions['I'].width = 13
        ws.column_dimensions['J'].width = 13


        lh = ['Batch ID','RT 015/03 Number', 'Status', 'Wholesaler','Creation Date','Block Date', 'Release Date', 'State', 'Charge Area'   ]
        j = 1
        for h in lh:
            ws.cell(row=1, column=j).value = h
            j += 1

        i = 2
        q = """
            select batchid, callerid, status, wskey, assigndate, blockdate, releasedate, code_area, state from GW_callerid
            where status!=0 and wskey = ? 
            order by assigndate 
            
            """
        # db.cur.execute(q, '%{0}%'.format(batchid))
        # rows = db.cur.fetchall()

        rows = db.cur.execute(q, wskey).fetchall()
        for r in rows:
            ws.cell(row=i, column=1).value = r.batchid
            ws.cell(row=i, column=2).value = r.callerid
            ws.cell(row=i, column=3).value = r.status
            ws.cell(row=i, column=4).value = r.wskey
            ws.cell(row=i, column=5).value = r.assigndate
            ws.cell(row=i, column=6).value = r.blockdate
            ws.cell(row=i, column=7).value = r.releasedate
            ws.cell(row=i, column=8).value = r.code_area
            ws.cell(row=i, column=9).value = r.state
            i += 1

        b = save_virtual_workbook(wb)

    finally:
        if db is not None:
            db.dispose()

    return b
    
def view015numbers_excel_list_batchid(batchid):
    
    print(batchid)
    b = None
    db = None

    try:
        db = initdb()

        wb = Workbook()
        ws = wb.create_sheet('Available Number', 0)

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 22
        ws.column_dimensions['H'].width = 13
        ws.column_dimensions['I'].width = 13
        ws.column_dimensions['J'].width = 13


        lh = ['Batch ID','RT 015/03 Number', 'Status', 'Wholesaler','Creation Date','Block Date', 'Release Date', 'State', 'Charge Area'   ]
        j = 1
        for h in lh:
            ws.cell(row=1, column=j).value = h
            j += 1

        i = 2
        q = """
            select batchid, callerid, status, wskey, assigndate, blockdate, releasedate, code_area, state from GW_callerid
            where status=0 and batchid like ? 
            order by assigndate 
        
            """
        db.cur.execute(q, '%{0}%'.format(batchid))
        rows = db.cur.fetchall()

        # rows = db.cur.execute(q, batchid).fetchall()
        for r in rows:
            ws.cell(row=i, column=1).value = r.batchid
            ws.cell(row=i, column=2).value = r.callerid
            ws.cell(row=i, column=3).value = r.status
            ws.cell(row=i, column=4).value = r.wskey
            ws.cell(row=i, column=5).value = r.assigndate
            ws.cell(row=i, column=6).value = r.blockdate
            ws.cell(row=i, column=7).value = r.releasedate
            ws.cell(row=i, column=8).value = r.code_area
            ws.cell(row=i, column=9).value = r.state
   
            i += 1

        ws = wb.create_sheet('Used Number', 1)

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 22
        ws.column_dimensions['H'].width = 13
        ws.column_dimensions['I'].width = 13
        ws.column_dimensions['J'].width = 13


        lh = ['Batch ID','RT 015/03 Number', 'Status', 'Wholesaler','Creation Date','Block Date', 'Release Date', 'State', 'Charge Area'   ]
        j = 1
        for h in lh:
            ws.cell(row=1, column=j).value = h
            j += 1

        i = 2
        q = """
            select batchid, callerid, status, wskey, assigndate, blockdate, releasedate, code_area, state from GW_callerid
            where status!=0 and batchid like ? 
            order by assigndate 
            
            """
        db.cur.execute(q, '%{0}%'.format(batchid))
        rows = db.cur.fetchall()

        # rows = db.cur.execute(q, batchid).fetchall()
        for r in rows:
            ws.cell(row=i, column=1).value = r.batchid
            ws.cell(row=i, column=2).value = r.callerid
            ws.cell(row=i, column=3).value = r.status
            ws.cell(row=i, column=4).value = r.wskey
            ws.cell(row=i, column=5).value = r.assigndate
            ws.cell(row=i, column=6).value = r.blockdate
            ws.cell(row=i, column=7).value = r.releasedate
            ws.cell(row=i, column=8).value = r.code_area
            ws.cell(row=i, column=9).value = r.state
            i += 1

        b = save_virtual_workbook(wb)

    finally:
        if db is not None:
            db.dispose()

    return b


def available015numbers_count(sfrom, sto, dbx=None):
    x = 0
    db = dbx

    try:
        db = initdb(dbx)

        q = """
            select count(callerid) as cnt from gw_callerid 
            where wskey in (0) and status = 0 
            and callerid between ? and ?
            """
        r = db.cur.execute(q, sfrom, sto).fetchone()
        if r is not None:
            x = r.cnt

    finally:
        if db is not None and dbx is None:
            db.dispose()

    return x

def used015numbers_excel_list(sfrom, sto):
    b = None
    db = None

    try:
        fromstr = sfrom.strip()
        tostr = sto.strip()

        db = initdb()

        wb = Workbook()
        ws = wb.create_sheet('Sheet1', 0)

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['C'].width = 11
        lh = ['RT 015/03 Number', 'Status', 'Wholesaler']
        j = 1
        for h in lh:
            ws.cell(row=1, column=j).value = h
            j += 1

        i = 2
        q = """
            select bb_rt015, bb_status, bb_wskey from bb_authentication 
            where bb_rt015 between ? and ?
            """
        rows = db.cur.execute(q, fromstr, tostr).fetchall()
        for r in rows:
            ws.cell(row=i, column=1).value = r.bb_rt015
            ws.cell(row=i, column=2).value = r.bb_status
            ws.cell(row=i, column=3).value = r.bb_wskey
            i += 1

        ws = wb.create_sheet('Sheet2', 1)

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 13
        ws.column_dimensions['C'].width = 25
        lh = ['RT 015/03 Number', 'Account ID', 'Wholesaler']
        j = 1
        for h in lh:
            ws.cell(row=1, column=j).value = h
            j += 1

        i = 2
        q = """
            select a.callerid, a.accountid, c.wholesalername 
            from authentication a left join userdetailext b on a.accountid = b.accountid
            left join twholesaler c on b.wholesalerkey = c.wholesalerkey
            where a.callerid between ? and ?
            """
        rows = db.cur.execute(q, fromstr, tostr).fetchall()
        for r in rows:
            ws.cell(row=i, column=1).value = r.callerid
            ws.cell(row=i, column=2).value = r.accountid
            ws.cell(row=i, column=3).value = r.wholesalername
            i += 1

        b = save_virtual_workbook(wb)

    finally:
        if db is not None:
            db.dispose()

    return b

def used015numbers_count(sfrom, sto, dbx=None):
    x = 0
    db = dbx

    try:
        db = initdb(dbx)

        q = """
            select count(callerid) as cnt from gw_callerid
            where wskey<> 0 and callerid between ? and ?
            """
        r = db.cur.execute(q, sfrom, sto).fetchone()
        if r is not None:
            x = r.cnt

    finally:
        if db is not None and dbx is None:
            db.dispose()

    return x

def assign015(sfrom, sto, expectedcount, wskey, user, pwd):
    x = 0
    y = 0
    db = None

    try:
        fromstr = sfrom.strip()
        tostr = sto.strip()

        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)

        x = available015numbers_count(fromstr, tostr, db)
        if int(x) != int(expectedcount):
            y = used015numbers_count(fromstr, tostr, db)
            if y > 0:
                raise UIException('{0} numbers are currently in use'.format(y))

            elif x > 0:
                raise UIException('Numbers count does not match')

        if x == 0 and y == 0:
            num1 = int(fromstr)
            num2 = int(tostr)
            n = 0
            
            if n == expectedcount:
                x = expectedcount
                db.commit()

            else:
                raise Exception('Failed to assign numbers {0} - {1} for Wholesaler {2}'.format(fromstr, tostr, wskey))

        elif int(x) == int(expectedcount):
            q = """
                update gw_callerid set wskey = ?, assignee = ?, assigndate = ?
                where wskey in (0) and status = 0 
                and callerid between ? and ?
                """
            db.cur.execute(q, wskey, user, datetime.now(), fromstr, tostr)
            db.commit()

    finally:
        if db is not None:
            db.dispose()

    return x

# def unassign015(sfrom, sto, user, pwd):
#     db = None

#     try:
#         fromstr = sfrom.strip()
#         tostr = sto.strip()

#         db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)

#         q = """
#             update bb_authentication set bb_allowpstn = 0, bb_status = 0, bb_wskey = 0 
#             where bb_rt015 between ? and ?
#             """
#         db.cur.execute(q, fromstr, tostr)
#         db.commit()

#     finally:
#         if db is not None:
#             db.dispose()

def unassign_wskey(sfrom, sto, user, pwd):
    db = None

    try:
        fromstr = sfrom.strip()
        tostr = sto.strip()

        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)

        q = """
            update gw_callerid set bb_wskey = 0 
            where callerid between ? and ?
            """
        db.cur.execute(q, fromstr, tostr)
       
        db.commit()

    finally:
        if db is not None:
            db.dispose()

def prepaid_excel_list():
    b = None
    db = None

    try:
        db = initdb()

        wb = Workbook()
        ws = wb.create_sheet('Sheet1', 0)

        ws.column_dimensions['A'].width = 18
        lh = ['RT 015/03 Number']
        j = 1
        for h in lh:
            ws.cell(row=1, column=j).value = h
            j += 1

        i = 2
        q = """
            select bb_rt015 from bb_authentication 
            where bb_rt015 in (
                select username from openquery(sipserver, 'select username, grp from grp')
                where grp = 'prepaid'
            ) 
            order by bb_rt015
            """
        rows = db.cur.execute(q).fetchall()
        for r in rows:
            ws.cell(row=i, column=1).value = r.bb_rt015
            i += 1

        b = save_virtual_workbook(wb)

    finally:
        if db is not None:
            db.dispose()

    return b


def download_used_list():
    b = None
    db = None

    try:
        db = initdb()

        wb = Workbook()
        ws = wb.create_sheet('Sheet1', 0)

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 22
        ws.column_dimensions['H'].width = 13
        ws.column_dimensions['I'].width = 13
        ws.column_dimensions['J'].width = 13
        lh = ['Batch ID','RT 015/03 Number', 'Status', 'Wholesaler','Creation Date','Block Date', 'Release Date', 'State', 'Charge Area'   ]
        j = 1
        for h in lh:
            ws.cell(row=1, column=j).value = h
            j += 1

        i = 2
        q = """
            select batchid, callerid, status, wskey, assigndate, blockdate, releasedate, code_area, state from GW_callerid
            where status!=0  
            order by assigndate 
        
            """
        
        rows = db.cur.execute(q).fetchall()


        # rows = db.cur.execute(q, batchid).fetchall()
        for r in rows:
            ws.cell(row=i, column=1).value = r.batchid
            ws.cell(row=i, column=2).value = r.callerid
            ws.cell(row=i, column=3).value = r.status
            ws.cell(row=i, column=4).value = r.wskey
            ws.cell(row=i, column=5).value = r.assigndate
            ws.cell(row=i, column=6).value = r.blockdate
            ws.cell(row=i, column=7).value = r.releasedate
            ws.cell(row=i, column=8).value = r.code_area
            ws.cell(row=i, column=9).value = r.state
   
            i += 1

        b = save_virtual_workbook(wb)

    finally:
        if db is not None:
            db.dispose()

    return b

def download_avail_list():

    b = None
    db = None

    try:
        db = initdb()

        wb = Workbook()
        ws = wb.create_sheet('Sheet1', 0)

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 22
        ws.column_dimensions['H'].width = 13
        ws.column_dimensions['I'].width = 13
        ws.column_dimensions['J'].width = 13
        lh = ['Batch ID','RT 015/03 Number', 'Status', 'Wholesaler','Creation Date','Block Date', 'Release Date', 'State', 'Charge Area'   ]
        j = 1
        for h in lh:
            ws.cell(row=1, column=j).value = h
            j += 1

        i = 2
        q = """
            select batchid, callerid, status, wskey, assigndate, blockdate, releasedate, code_area, state from GW_callerid
            where status =0  
            order by assigndate 
        
            """
        
        rows = db.cur.execute(q).fetchall()
        # rows = db.cur.execute(q, batchid).fetchall()
        for r in rows:
            ws.cell(row=i, column=1).value = r.batchid
            ws.cell(row=i, column=2).value = r.callerid
            ws.cell(row=i, column=3).value = r.status
            ws.cell(row=i, column=4).value = r.wskey
            ws.cell(row=i, column=5).value = r.assigndate
            ws.cell(row=i, column=6).value = r.blockdate
            ws.cell(row=i, column=7).value = r.releasedate
            ws.cell(row=i, column=8).value = r.code_area
            ws.cell(row=i, column=9).value = r.state
   
            i += 1

        b = save_virtual_workbook(wb)

    finally:
        if db is not None:
            db.dispose()

    return b

def wslogin_create(o):
    db = None

    try:
        assert isinstance(o, models.WSLogin)
        db = initdb()
        ws = iswslogin_exist(o.wsid, db)
        if ws is not None:
            raise UIException('Login for wholesaler {0} - {1} already exist'.format(o.wsid, o.wsname))

        q = """
            if not exists(select 1 from registrationConfig where wholesalerid = ?)
            begin
              insert into registrationConfig 
              select ?, ?, DataType, DataValue, ? from registrationConfig where wholesalerid = 'BZGroup'

              update registrationConfig set DataValue = ? where wholesalerid = ? and DataType = 'WSKEY'
            end
            """
        db.cur.execute(q, o.loginid, o.loginid, o.loginid, o.pattern, o.wsid, o.loginid)

        q = """
            SET NOCOUNT ON
            declare
              @return_value int,
		      @ErrorCode int
            exec @return_value = spp_WebUser_New
              @UserLoginID = ?,
              @NewPassword = ?,
              @WholesalerKey = ?,
              @ErrorCode = @ErrorCode output
            select
              @ErrorCode as N'ErrorCode',
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, o.loginid, o.pwd, o.loginid).fetchone()
        if geterrorcode(r) == 1:
            raise UIException('Invalid Password (length < 6 digits)')

        elif geterrorcode(r) == 2:
            raise UIException('Cannot find Agent account to set password')

    finally:
        if db is not None:
            db.dispose()

def rtcdruser_create(o):
    db = None

    try:
        assert isinstance(o, models.WSLogin)
        db = initdb()
        k = isrtcdruser_exist(o.wsid, db)
        if k is not None:
            raise UIException('Login for wholesaler {0} - {1} already exist'.format(o.wsid, o.wsname))

        q = """
            SET NOCOUNT ON
            declare	
              @return_value int,
		      @ErrorCode int
            exec @return_value = RTCDR_sp_User_New
              @UserLoginID = ?,
		      @UserPassword = ?,
		      @UserLevel = 1,
		      @WholesalerID = ?,
		      @CustIDFilter = ?,
		      @ErrorCode = @ErrorCode output
            select
              @ErrorCode as N'ErrorCode',
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, o.loginid, o.pwd, str(o.wsid), o.pattern).fetchone()
        if geterrorcode(r) != 0:
            raise UIException('Failed to create login')

    finally:
        if db is not None:
            db.dispose()

def iswslogin_exist(wskey, dbx=None):
    ws = None
    db = dbx

    try:
        db = initdb(dbx)

        q = """
            select wholesalerid from registrationconfig 
            where datatype = 'WSKEY' and datavalue = ?
            """
        r = db.cur.execute(q, str(wskey)).fetchone()
        if r is not None:
            ws = r.wholesalerid

    finally:
        if db is not None and dbx is None:
            db.dispose()

    return ws

def isrtcdruser_exist(wskey, dbx=None):
    k = None
    db = dbx

    try:
        db = initdb(dbx)

        q = """
            select loginid from rtcdr_user 
            where wholesalerid = ?
            """
        r = db.cur.execute(q, str(wskey)).fetchone()
        if r is not None:
            k = r.loginid

    finally:
        if db is not None and dbx is None:
            db.dispose()

    return k

def list_wslogin(dbx=None):
    l = []
    db = dbx

    try:
        db = initdb(dbx)

        q = """
            select sloginname, wholesalerkey from twholesalerlogin 
            where wholesalerkey <> ''
            order by sloginname
            """
        rows = db.cur.execute(q).fetchall()
        for r in rows:
            o = models.WholesalerLogin()
            o.sloginname = r.sloginname
            o.wholesalerkey = r.wholesalerkey
            l.append(o)

    finally:
        if db is not None and dbx is None:
            db.dispose()

    return l

def list_rtcdruser(dbx=None):
    l = []
    db = dbx

    try:
        db = initdb(dbx)

        q = """
            select loginid, wholesalerid from rtcdr_user 
            where wholesalerid <> '*' 
            order by loginid
            """
        rows = db.cur.execute(q).fetchall()
        for r in rows:
            o = models.RTCDRUser()
            o.loginid = r.loginid
            o.wholesalerid = r.wholesalerid
            l.append(o)

    finally:
        if db is not None and dbx is None:
            db.dispose()

    return l

def wspassword_get(s, wskey):
    ret = s
    db = None

    try:
        db = initdb()

        q = """
            select wholesalerid, datavalue from registrationconfig 
            where wholesalerid = ? and datatype = 'WSKEY'
            """
        r = db.cur.execute(q, wskey).fetchone()
        if r is not None:
            k = r.datavalue
            j = k.zfill(4)
            ret = '{0}{1}'.format(s, j)

    finally:
        if db is not None:
            db.dispose()

    return ret

def rtcdruserpassword_get(s, wsid):
    j = wsid.zfill(4)
    ret = '{0}{1}'.format(s, j)
    return ret

def wspassword_update(login, pwd):
    db = None

    try:
        db = initdb()

        q = """
            SET NOCOUNT ON
            declare 
              @return_value int,
              @ErrorCode int
            exec @return_value = spp_WebUser_SetPassword
              @UserLoginID = ?,
		      @NewPassword = ?,
		      @ErrorCode = @ErrorCode output
            select
              @ErrorCode as N'ErrorCode',
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, login, pwd).fetchone()
        if geterrorcode(r) == 1:
            raise UIException('Invalid Password (length < 6 digits)')

        elif geterrorcode(r) == 2:
            raise UIException('Cannot find Agent account to set password')
    
    finally:
        if db is not None:
            db.dispose()

def rtcdruserpassword_update(login, pwd):
    db = None

    try:
        db = initdb()

        q = """
            SET NOCOUNT ON
            declare 
              @return_value int,
              @ErrorCode int
            exec @return_value = RTCDR_sp_User_ChangePassword
              @UserLoginID = ?,
		      @UserPassword = ?,
		      @ErrorCode = @ErrorCode output
            select
              @ErrorCode as N'ErrorCode',
              @return_value as N'Ret'
            """
        r = db.cur.execute(q, login, pwd).fetchone()
        if geterrorcode(r) != 0:
            raise UIException('Failed to set password')
    
    finally:
        if db is not None:
            db.dispose()

def custid_pattern_get(wskey):
    x = None
    db = None

    try:
        db = initdb()

        q = """
            select top 1 idmask from registrationconfig 
            where datatype = 'WSKEY' and datavalue = ?
            """
        r = db.cur.execute(q, wskey).fetchone()
        if r is not None:
            x = r.idmask

        if x is None:
            q = """
                select custidfilter from rtcdr_user 
                where wholesalerid = ?
                """
            r = db.cur.execute(q, wskey).fetchone()
            if r is not None:
                x = r.custidfilter

    finally:
        if db is not None:
            db.dispose()

    return x

def wholesaler_get(accountid):
    wskey = None
    db = None

    try:
        db = initdb()
        
        q = 'select wholesalerkey from userdetailext where accountid = ?'
        r = db.cur.execute(q, accountid).fetchone()
        if r is not None:
            wskey = r.wholesalerkey

    finally:
        if db is not None:
            db.dispose()

    return wskey

def geterrorcode(r):
    x = 0
    try:
        x = r.ErrorCode

    except:
        pass

    return x


def createbatch():
    db = None
    l = []

    try:
        
        db = initdb()
        q = """
            select top 12 custid, amount, realdate from customersettlement
            where custid = ? and (paymenttype = 1 or (paymenttype = 4 and productid = 6))
            order by settlementidx desc
            """
        rows = db.cur.execute(q, custid).fetchall()
        for r in rows:
            m = utils.sqltodic(r, db)
            m['date'] = utils.fmtdate(r.realdate)
            l.append(m)

    finally:
        disposedb(db)

    return l

def upload015(batch_id, sfrom, sto, expectedcount,  user, pwd):
    x = 0
    y = 0
    db = None
    
    try:
        fromstr = sfrom.strip()
        tostr = sto.strip()
        print(fromstr)
        print(tostr)


        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        # bc=batch_check(batch_id, db)
        # if bc > 0:
        #     raise UIException('BatchID name currently exist'.format(y))
     

        x = available015upload_count(fromstr, tostr, db)
        if x != expectedcount:
            y = used015_count(fromstr, tostr, db)
            if y > 0:
                raise UIException('{0} numbers are currently in use'.format(y))

            elif x > 0:
                raise UIException('Numbers count does not match')

        if x == 0 and y == 0:
            print(fromstr)
            num1 = int(fromstr)
            # print(num1)
            num2 = int(tostr)
            n = 0
            print(batch_id)
            for i in range(num1, num2 + 1):
                num = str(i).zfill(len(fromstr))
                print(num)


                db = initdb()
                q = """
                    INSERT into GW_callerid (batchid, callerid, assigndate, status, wskey, assignee)
                    values (?, ?, current_timestamp, 0, 0,?)
                    """
                print(q)
                print(num)
                db.cur.execute(q, batch_id.upper(), num, user)
                n += 1
                print(n)
                print(expectedcount)

            if n == expectedcount:
                x = expectedcount
                db.commit()
                

            # else:
                # raise Exception('Failed to upload numbers {0} - {1} for Wholesaler {2}'.format(fromstr, tostr, batch_id))

        # elif x == expectedcount:
        #     q = """
        #         update bb_authentication set bb_allowpstn = 1, bb_wskey = ? 
        #         where bb_wskey in (0, 1, 73) and bb_status = 0 
        #         and bb_rt015 between ? and ?
        #         """
        #     db.cur.execute(q, wskey, fromstr, tostr)
        #     db.commit()

    finally:
        if db is not None:
            db.dispose()

    return x


def batchinfo(batch_id,expectedcount,user):
    y = 0

    try:
        
        db = initdb()

        q = """
            insert into GW_BatchNumber(batch_id, uploadby, batch_date, batch_qty)
            values (?, ?, current_timestamp, ?)
            """
        print(q)
        db.cur.execute(q, batch_id.upper(), user, expectedcount)
        db.commit()

        

    finally:
        if db is not None and dbx is None:
            db.dispose()

    return y

def available015upload_count(sfrom, sto, dbx=None):
    x = 0
    db = dbx

    try:
        db = initdb(dbx)

        q = """
            select count(callerid) as cnt from GW_callerid 
            where status = 0 
            and callerid between ? and ?
            """
        r = db.cur.execute(q, sfrom, sto).fetchone()
        if r is not None:
            x = r.cnt

    finally:
        if db is not None and dbx is None:
            db.dispose()

    return x

def used015_count(sfrom, sto, dbx=None):
    x = 0
    db = dbx

    try:
        db = initdb(dbx)

        q = """
            select count(callerid) as cnt from GW_callerid
            where callerid between ? and ?
            """
        r = db.cur.execute(q, sfrom, sto).fetchone()
        if r is not None:
            x = r.cnt

    finally:
        if db is not None and dbx is None:
            db.dispose()

    return x

def usedstate_count(sfrom, sto, dbx=None):
    x = 0
    db = dbx

    try:
        db = initdb(dbx)

        q = """
            select count(callerid) as cnt from GW_callerid
            where (state is not null or code_area is not null) and callerid between ? and ?
            """
        r = db.cur.execute(q, sfrom, sto).fetchone()
        if r is not None:
            x = r.cnt

    finally:
        if db is not None and dbx is None:
            db.dispose()

    return x

def assign_state(code_area,state, sfrom, sto, expectedcount,  user, pwd):
    x = 0
    y = 0
    db = None

    try:
        fromstr = sfrom.strip()
        tostr = sto.strip()

        db = initdb(dbx=None, autocommit=False, user=user, pwd=pwd)
        x = available015assignstate_count(fromstr, tostr, db)
        if int(x) != int(expectedcount):
            y = usedstate_count(fromstr, tostr, db)
            if y > 0:
                raise UIException('{0} numbers are currently in use'.format(y))

            elif x > 0:
                raise UIException('Numbers count does not match')

        
        elif int(x) == int(expectedcount):


            q = """
                UPDATE GW_callerid
                SET state = ?, code_area = ?
                WHERE callerid between ? and ?
                
                """
            db.cur.execute(q, code_area, state, fromstr, tostr)
            db.commit()

    finally:
        if db is not None:
            db.dispose()

    return x


def available015assignstate_count(sfrom, sto, dbx=None):
    x = 0
    db = dbx

    try:
        db = initdb(dbx)

        q = """
            select count(callerid) as cnt from GW_callerid 
            where status = 0 and state is null and code_area is null
            and callerid between ? and ?
            """
        print(q)

        r = db.cur.execute(q, sfrom, sto).fetchone()
        if r is not None:
            x = r.cnt

    finally:
        if db is not None and dbx is None:
            db.dispose()

    return x

def batchidlist(dbx=None):
    l = []
    db = None
    
    try:
        db = initdb()
        
        q = "select batch_id, batch_qty from GW_BatchNumber order by batch_date desc"
        print(q)
        rows = db.cur.execute(q).fetchall()
        for r in rows:
            o = models.batchID()
            o.batch_id = r.batch_id
            o.batch_qty = r.batch_qty
            l.append(o)
            # print(l)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def list_area(dbx=None):
    l = []
    db = None
    
    try:
        db = initdb()
        
        q = "select codearea, statearea from gw_statecode order by id desc"
        print(q)
        rows = db.cur.execute(q).fetchall()
        for r in rows:
            o = models.Codearea()
            o.codearea = r.codearea
            o.statearea = utils.decode(r.statearea)
            # print(o.codearea)
            l.append(o)
            # print(l)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return l

def list_subarea(dbx=None):
    ls = []
    db = None
    
    try:
        db = initdb()
        
        q = "select codearea, charearea from gw_state order by id desc"
        print(q)
        rows = db.cur.execute(q).fetchall()
        for r in rows:
            o = models.statearea()
            o.codearea = r.codearea
            o.charearea = utils.decode(r.charearea)
            # print(o.codearea)
            ls.append(o)
            # print(l)
    
    finally:
        if db is not None and dbx is None:
            db.dispose()
            
    return ls

def chargearea_add(codearea,charearea,user):
    db = None

    try:
        
        db = initdb()

        q = """
            insert into gw_state(codearea, charearea, createdby)
            values (?, ?, ?)
            """
        print(codearea)
        print(user)

        db.cur.execute(q, str(codearea), str(charearea), user)
        db.commit()

        

    finally:
        if db is not None:
            db.dispose()

    # return y


def insert_into_deleted_record(batchid, callerid, wskey, code_area,status, user): 
    db = None
    try:
        db = initdb()

        q = """
            insert into GW_deleted_callerid (batchid, callerid, date, wskey, status, code_area, [user]) 
            values (?, ?, ?, ?, ?,?,?)
            """
        print(q)
        db.cur.execute(q,batchid, callerid,  datetime.now(), wskey, status, code_area, user)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()

def insert_into_blocked_record(batchid, callerid, wskey, code_area,status, user): 
    db = None
    try:
        db = initdb()

        q = """
            insert into GW_blocked_log (batchid, callerid, date, wskey, status, code_area, [user]) 
            values (?, ?, ?, ?, ?,?,?)
            """
        print(q)
        db.cur.execute(q,batchid, callerid,  datetime.now(), wskey, status, code_area, user)
        db.commit()
        
    finally:
        if db is not None:
            db.dispose()